import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from plotly.colors import sample_colorscale
import pydeck as pdk
import re
from pathlib import Path
from io import BytesIO
import time
import importlib
import os
import json
import math
from typing import Any, Dict, Optional, Tuple

# ── Dark Plotly theme ─────────────────────────────────────────────────────────
_PLT_BG   = "#060d18"
_PLT_PAPER= "#0a1525"
_PLT_GRID = "rgba(6,214,240,.06)"
_PLT_LINE = "rgba(6,214,240,.18)"
_PLT_TICK = "#5aa8c4"
_PLT_TITLE= "#dff2fa"
_PLT_LEGEND_BG = "rgba(10,21,37,.9)"
_PLT_LEGEND_BORDER = "rgba(6,214,240,.2)"
_PLT_HOVER_BG = "#0a1525"
_PLT_HOVER_FONT = "#dff2fa"
_PLT_SERIES = ["#ef4444", "#2563eb", "#16a34a", "#9333ea", "#ea580c", "#ca8a04", "#0891b2", "#db2777", "#6b7280", "#0f766e"]
_PLT_MARKER_BORDER = "#f8fafc"

def _dark_layout(**extra):
    base = dict(
        plot_bgcolor=_PLT_BG,
        paper_bgcolor=_PLT_PAPER,
        font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
        xaxis=dict(showgrid=True, gridcolor=_PLT_GRID, linecolor=_PLT_LINE,
                   tickfont=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"), title_font=dict(color=_PLT_TITLE)),
        yaxis=dict(showgrid=True, gridcolor=_PLT_GRID, linecolor=_PLT_LINE,
                   tickfont=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"), title_font=dict(color=_PLT_TITLE)),
        legend=dict(bgcolor=_PLT_LEGEND_BG, bordercolor=_PLT_LEGEND_BORDER,
                    borderwidth=1, font=dict(color=_PLT_TICK)),
        hoverlabel=dict(bgcolor=_PLT_HOVER_BG, font_color=_PLT_HOVER_FONT, bordercolor=_PLT_LEGEND_BORDER),
    )
    base.update(extra)
    return base

def _use_light_plot_theme():
    global _PLT_BG, _PLT_PAPER, _PLT_GRID, _PLT_LINE, _PLT_TICK, _PLT_TITLE
    global _PLT_LEGEND_BG, _PLT_LEGEND_BORDER, _PLT_HOVER_BG, _PLT_HOVER_FONT, _PLT_MARKER_BORDER
    _PLT_BG = "#ffffff"
    _PLT_PAPER = "#ffffff"
    _PLT_GRID = "rgba(15,23,42,.10)"
    _PLT_LINE = "rgba(15,23,42,.18)"
    _PLT_TICK = "#334155"
    _PLT_TITLE = "#0f172a"
    _PLT_LEGEND_BG = "rgba(255,255,255,.96)"
    _PLT_LEGEND_BORDER = "rgba(15,23,42,.12)"
    _PLT_HOVER_BG = "#ffffff"
    _PLT_HOVER_FONT = "#0f172a"
    _PLT_MARKER_BORDER = "#0f172a"

def _apply_dark(fig, height=None, title=None, **kw):
    layout = _dark_layout(**kw)
    if height:
        layout["height"] = height
    if title:
        layout["title"] = dict(text=title, font=dict(color=_PLT_TITLE, size=14))
    fig.update_layout(**layout)
    fig.update_xaxes(showgrid=True, gridcolor=_PLT_GRID, linecolor=_PLT_LINE,
                     tickfont=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"), title_font=dict(color=_PLT_TITLE))
    fig.update_yaxes(showgrid=True, gridcolor=_PLT_GRID, linecolor=_PLT_LINE,
                     tickfont=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"), title_font=dict(color=_PLT_TITLE))
    return fig

try:
    import ee  # type: ignore
except Exception:
    ee = None

try:
    import folium  # type: ignore
    from folium.plugins import Draw  # type: ignore
    from streamlit_folium import st_folium  # type: ignore
except Exception:
    folium = None
    Draw = None
    st_folium = None

st.set_page_config(
    page_title="Δορυφορική Παρακολούθηση Φράγματος Γαδουρά",
    page_icon="💧",
    layout="wide",
    initial_sidebar_state="expanded"
)

APP_DIR = Path(__file__).resolve().parent
def _resolve_platform_root(app_dir: Path) -> Path:
    # Supports both layouts:
    # 1) script at repo root with sibling data folders
    # 2) script under `code/` with data folders at parent
    direct_has_data = (app_dir / "satellite data").exists() or (app_dir / "field data").exists()
    parent_has_data = (app_dir.parent / "satellite data").exists() or (app_dir.parent / "field data").exists()
    if direct_has_data:
        return app_dir
    if parent_has_data:
        return app_dir.parent
    return app_dir

PLATFORM_ROOT = _resolve_platform_root(APP_DIR)
FIELD_DATA_ROOT = PLATFORM_ROOT / "field data"
SATELLITE_DATA_ROOT = PLATFORM_ROOT / "satellite data"
DELIVERABLES_ROOT = PLATFORM_ROOT / "Παραδοτέα"
SHARED_DATA_ROOT = SATELLITE_DATA_ROOT / "DATA"
if not SHARED_DATA_ROOT.exists():
    legacy_data_root = PLATFORM_ROOT / "DATA"
    if legacy_data_root.exists():
        SHARED_DATA_ROOT = legacy_data_root

APP_HEADER_TITLE = "Εφαρμογή Ανάλυσης Ποιότητας Επιφανειακών Υδάτων για τον Ταμιευτήρα Γαδουρά ΕΥΑΘ ΑΕ"
APP_HEADER_SUBTITLE = "Οπτικοποίηση δορυφορικών GeoTIFF και επικυρωμένων μετρήσεων"
APP_LOGO_URL = "https://chatbot.eyath.gr/_astro/eyath-logo-2.DriaSExn_1jOI34.svg"


def render_main_header() -> None:
    st.markdown(
        f"""<div class="hcard">
              <img src="{APP_LOGO_URL}" style="height:60px;object-fit:contain;flex-shrink:0;"
                   onerror="this.style.display='none'" />
              <div>
                <h1>{APP_HEADER_TITLE}</h1>
                <div class="sub">{APP_HEADER_SUBTITLE}</div>
                <span class="badge">💧 Gadouras Reservoir · Rhodes, GR</span>
              </div>
            </div>""",
        unsafe_allow_html=True,
    )


def render_satellite_data_view() -> None:
    try:
        satellite_app = importlib.import_module("streamlit_geotiff_map_1")
    except Exception as exc:
        st.error("Αποτυχία φόρτωσης της ενότητας δορυφορικών δεδομένων.")
        st.caption(f"Module import error: {exc}")
        return

    render_fn = getattr(satellite_app, "render_satellite_dashboard", None)
    if not callable(render_fn):
        st.error("Το module `streamlit_geotiff_map_1.py` δεν εκθέτει τη συνάρτηση `render_satellite_dashboard`.")
        return

    st.caption(f"Storage root: `{PLATFORM_ROOT}`")
    st.caption(f"Satellite source root: `{SATELLITE_DATA_ROOT}`")
    try:
        # Render inline in the same Streamlit process (Cloud-safe; no localhost iframe).
        render_fn(show_header=False, show_footer=False, show_debug=False, apply_css=False)
    except Exception as exc:
        st.error("Αποτυχία προβολής δορυφορικών δεδομένων.")
        st.caption(f"Render error: {exc}")


def _deliverable_sort_key(path: Path):
    match = re.search(r"(\d+)", path.stem)
    return (int(match.group(1)) if match else 10_000, path.name.lower())


@st.cache_data(show_spinner=False)
def _read_binary_file(path: str) -> bytes:
    return Path(path).read_bytes()


def render_deliverables_view() -> None:
    st.subheader("Παραδοτέα Έργου")
    st.caption(f"Φάκελος παραδοτέων: `{DELIVERABLES_ROOT}`")

    if not DELIVERABLES_ROOT.exists():
        st.error(f"Δεν βρέθηκε ο φάκελος παραδοτέων: `{DELIVERABLES_ROOT}`")
        return

    deliverable_files = sorted(
        [path for path in DELIVERABLES_ROOT.iterdir() if path.is_file()],
        key=_deliverable_sort_key,
    )
    if not deliverable_files:
        st.info("Δεν βρέθηκαν διαθέσιμα παραδοτέα.")
        return

    selected_file = st.selectbox(
        "Επιλέξτε παραδοτέο",
        options=deliverable_files,
        format_func=lambda path: path.name,
        key="deliverable_file_selector",
    )

    file_bytes = _read_binary_file(str(selected_file))
    file_size_mb = len(file_bytes) / (1024 * 1024)

    info_col, meta_col, action_col = st.columns([3, 1.4, 1.6])
    with info_col:
        st.markdown(f"### {selected_file.stem}")
    with meta_col:
        st.caption(f"Μέγεθος: {file_size_mb:.2f} MB")
    with action_col:
        st.download_button(
            "Λήψη αρχείου",
            data=file_bytes,
            file_name=selected_file.name,
            mime="application/pdf" if selected_file.suffix.lower() == ".pdf" else "application/octet-stream",
            use_container_width=True,
            key=f"download_{selected_file.name}",
        )

    suffix = selected_file.suffix.lower()
    if suffix == ".pdf":
        st.info("Για τα PDF χρησιμοποιήστε τη λήψη αρχείου.")
    elif suffix in {".png", ".jpg", ".jpeg"}:
        st.image(file_bytes, use_container_width=True)
    else:
        st.info("Υποστηρίζεται preview μόνο για εικόνες. Για αυτό το αρχείο χρησιμοποιήστε λήψη.")

# ─── Sampling point coordinates (from docx) ───────────────────────────────────
SAMPLING_POINTS = {
    0: {"lat": 36.1586669, "lon": 27.994512},
    1: {"lat": 36.162806,  "lon": 27.997548},
    2: {"lat": 36.164875,  "lon": 28.002791},
    3: {"lat": 36.168398,  "lon": 27.996214},
    4: {"lat": 36.170233,  "lon": 27.985003},
    5: {"lat": 36.175543,  "lon": 27.981300},
    6: {"lat": 36.181249,  "lon": 27.977094},
    7: {"lat": 36.185767,  "lon": 27.973974},
    8: {"lat": 36.180363,  "lon": 27.969781},
}

# ─── Column mapping (0-based index in row) ────────────────────────────────────
COL_MAP = {
    "pH": 3,
    "Θερμοκρασία (°C)": 4,
    "Διαλυμένο Οξυγόνο DO (mg/L)": 5,
    "Αγωγιμότητα (μS/cm)": 6,
    "Θολότητα-Πεδίο (NTU)": 7,
    "Θολότητα-Εργαστήριο (NTU)": 8,
    "Χλωροφύλλη-α (μg/L)": 9,
    "Ολικά Αιωρούμενα Στερεά (mg/mL)": 10,
    "Δίσκος Secchi (m)": 11,
    "Ca (mg/L)": 12,
    "Mg": 13,
    "Αλκαλικότητα": 14,
    "NO3⁻ (mg/L)": 15,
    "NH4⁺ (mg/L)": 16,
    "PO4³⁻ (mg/L)": 17,
    "Fe²⁺ (mg/L)": 18,
    "Mn²⁺ (mg/L)": 19,
    "TOC (mg/L)": 20,
    "Ολικό Άζωτο TN": 21,
    "Ολικός Φώσφορος TP": 22,
}

DEPTH_ORDER = [
    "Επιφάνεια (5-10 cm)", "Βάθος 1m", "Βάθος 5m", "Βάθος 10m",
    "Βάθος 12m", "Βάθος 15m", "Βάθος 20m", "Βάθος 23m", "Βάθος 25m"
]

def try_float(v):
    if v is None:
        return np.nan
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    s = re.sub(r"^<\s*", "", s)  # remove leading "<"
    try:
        return float(s)
    except:
        return np.nan

def depth_to_m(label):
    if label is None:
        return np.nan
    s = str(label)
    if "5-10" in s:
        return 0.05
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*m", s)
    if m:
        return float(m.group(1).replace(",", "."))
    return np.nan


def _looks_like_depth_label(value):
    if value is None:
        return False
    s = str(value).strip().lower()
    if not s or s == "none":
        return False
    if "5-10" in s:
        return True
    if re.search(r"\d+(?:[.,]\d+)?\s*m", s):
        return True
    if "cm" in s and re.search(r"\d", s):
        return True
    return False

def _parse_rgb(color_text):
    nums = re.findall(r"\d+(?:\.\d+)?", str(color_text))
    if len(nums) >= 3:
        return [int(float(nums[0])), int(float(nums[1])), int(float(nums[2]))]
    return [56, 189, 248]

def _values_to_rgba(values, vmin, vmax, colorscale="Turbo", alpha=220):
    arr = pd.to_numeric(values, errors="coerce")
    if pd.isna(vmin) or pd.isna(vmax) or vmax <= vmin:
        return [[56, 189, 248, alpha] for _ in range(len(arr))]
    norm = ((arr - float(vmin)) / (float(vmax) - float(vmin))).clip(0, 1).fillna(0)
    out = []
    for n in norm:
        clr = sample_colorscale(colorscale, [float(n)])[0]
        rgb = _parse_rgb(clr)
        out.append([rgb[0], rgb[1], rgb[2], alpha])
    return out

@st.cache_data
def load_data(path):
    src_path = Path(path)
    last_err = None
    wb = None

    # OneDrive/Excel can briefly lock the workbook; retry and fallback to in-memory bytes.
    for attempt in range(6):
        try:
            wb = openpyxl.load_workbook(src_path, data_only=True)
            break
        except PermissionError as e:
            last_err = e
            try:
                raw = src_path.read_bytes()
                wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
                break
            except Exception as e_mem:
                last_err = e_mem
        except OSError as e:
            if getattr(e, "errno", None) == 13:
                last_err = e
            else:
                raise

        if attempt < 5:
            time.sleep(0.4 * (attempt + 1))

    if wb is None:
        if last_err is not None:
            raise last_err
        raise PermissionError(f"Cannot access workbook: {src_path}")

    records = []
    
    for sheet_name in wb.sheetnames:
        # Keep only date-like sheets (e.g. "10.09.2025"), skip template/meta sheets.
        try:
            date = pd.to_datetime(str(sheet_name).strip(), format="%d.%m.%Y", errors="raise")
        except Exception:
            continue
        
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        
        current_point = None
        
        for i, row in enumerate(rows):
            cell1 = ""
            if len(row) > 1 and row[1] is not None:
                cell1 = str(row[1]).strip()

            # Encoding-agnostic point-header detection:
            # match rows like "...: 0", "...: 4", etc.
            # Avoids false positives from volume/date rows such as "ΟΓΚΟΣ: 11,5 L *"
            # or "Ημερομηνία δειγματοληψίας: 10/09/2025".
            cell1_norm = cell1.lower()
            point_match = re.search(r":\s*(\d{1,2})\s*$", cell1_norm)
            if point_match:
                point_id = int(point_match.group(1))
                if 0 <= point_id <= 99:
                    current_point = point_id
                    continue
            
            # Data rows (structure-based, not header-text based)
            if current_point is not None and _looks_like_depth_label(cell1):
                depth = cell1.strip()
                rec = {
                    "date": date,
                    "point": current_point,
                    "depth": depth,
                }
                for param, col_idx in COL_MAP.items():
                    if col_idx < len(row):
                        rec[param] = try_float(row[col_idx])
                    else:
                        rec[param] = np.nan
                records.append(rec)
    
    df = pd.DataFrame(records)
    expected_cols = ["date", "point", "depth"] + list(COL_MAP.keys())
    if df.empty:
        return pd.DataFrame(columns=expected_cols)
    for col in expected_cols:
        if col not in df.columns:
            df[col] = np.nan
    if not df.empty:
        df = df.sort_values("date")
    return df


@st.cache_data
def load_level_data(root: str) -> pd.DataFrame:
    base = Path(root)
    if not base.exists():
        return pd.DataFrame(columns=["date", "value", "col", "source"])

    patterns = [
        "*level*.csv",
        "*Level*.csv",
        "*υψος*.csv",
        "*ύψος*.csv",
        "*stath*.csv",
        "*storage*.csv",
        "*water*.csv",
    ]
    candidates: list[Path] = []
    for pat in patterns:
        candidates.extend(sorted(base.rglob(pat)))
    if not candidates:
        candidates = sorted(base.rglob("*.csv"))

    # Deduplicate while preserving order.
    seen = set()
    unique_candidates: list[Path] = []
    for p in candidates:
        key = str(p.resolve()).lower()
        if key in seen:
            continue
        seen.add(key)
        unique_candidates.append(p)

    def _read_csv_any(path: Path) -> Optional[pd.DataFrame]:
        for enc in ("utf-8-sig", "utf-8", "cp1253", "latin-1"):
            try:
                return pd.read_csv(path, encoding=enc)
            except Exception:
                continue
        return None

    for csv_path in unique_candidates:
        raw = _read_csv_any(csv_path)
        if raw is None or raw.empty:
            continue

        cols = [str(c) for c in raw.columns]
        low = {c: c.lower() for c in cols}
        date_col = next(
            (c for c in cols if any(k in low[c] for k in ["date", "ημερ", "ημερο", "sample"])),
            cols[0],
        )
        val_col = next(
            (
                c
                for c in cols
                if any(k in low[c] for k in ["level", "υψ", "ύψ", "height", "storage", "value", "στάθ"])
            ),
            None,
        )
        if val_col is None:
            numeric_cols = [
                c for c in cols
                if pd.to_numeric(raw[c], errors="coerce").notna().sum() >= max(3, int(len(raw) * 0.25))
            ]
            if not numeric_cols:
                continue
            val_col = numeric_cols[0] if numeric_cols[0] != date_col else (numeric_cols[1] if len(numeric_cols) > 1 else numeric_cols[0])

        out = pd.DataFrame(
            {
                "date": pd.to_datetime(raw[date_col], errors="coerce"),
                "value": pd.to_numeric(raw[val_col], errors="coerce"),
                "col": val_col,
                "source": str(csv_path.resolve()),
            }
        ).dropna(subset=["date", "value"])
        if not out.empty:
            return out.sort_values("date").reset_index(drop=True)

    return pd.DataFrame(columns=["date", "value", "col", "source"])


def render_level_tab() -> None:
    st.subheader("Ύψος Στάθμης Ταμιευτήρα")
    st.caption(f"Φάκελος δεδομένων στάθμης: `{SHARED_DATA_ROOT}`")

    lvl = load_level_data(str(SHARED_DATA_ROOT))
    if lvl.empty:
        st.info(
            "Δεν βρέθηκε αρχείο CSV για την στάθμη. "
            f"Τοποθετήστε αρχείο με 'level' ή 'υψος' στο όνομα στον φάκελο: `{SHARED_DATA_ROOT}`"
        )
        return

    source_path = lvl["source"].iloc[0] if "source" in lvl.columns else "N/A"
    st.caption(f"Αρχείο στάθμης: `{source_path}`")
    val_lbl = lvl["col"].iloc[0] if "col" in lvl.columns else "Στάθμη (m)"
    dfp = lvl[["date", "value"]].copy()

    c1, c2, c3 = (st.columns(1) + st.columns(2)) if _is_mobile else st.columns([2, 1, 1])
    with c1:
        drng = st.date_input(
            "Εύρος ημερομηνιών",
            value=(dfp["date"].min().date(), dfp["date"].max().date()),
            min_value=dfp["date"].min().date(),
            max_value=dfp["date"].max().date(),
            key="lvl_main_range",
        )
    with c2:
        smooth = st.slider("Εξομάλυνση (ημέρες)", 1, 90, 1, key="lvl_main_smooth")
    with c3:
        ctype = st.radio("Τύπος", ["Εμβαδόν", "Γραμμή"], horizontal=True, key="lvl_main_type")

    if isinstance(drng, (list, tuple)) and len(drng) == 2:
        dfp = dfp[
            (dfp["date"] >= pd.Timestamp(drng[0])) &
            (dfp["date"] <= pd.Timestamp(drng[1]))
        ].copy()

    if dfp.empty:
        st.warning("Κανένα δεδομένο στο επιλεγμένο εύρος.")
        return

    dfp["display"] = dfp["value"].rolling(smooth, min_periods=1).mean()

    m1, m2, m3, m4 = (st.columns(2) + st.columns(2)) if _is_mobile else st.columns(4)
    m1.metric("Τελευταία", f"{dfp['value'].iloc[-1]:.2f} m")
    m2.metric("Μέγιστη", f"{dfp['value'].max():.2f} m")
    m3.metric("Ελάχιστη", f"{dfp['value'].min():.2f} m")
    m4.metric("Μέση", f"{dfp['value'].mean():.2f} m")

    fig = go.Figure()
    if ctype == "Εμβαδόν":
        fig.add_trace(
            go.Scatter(
                x=dfp["date"],
                y=dfp["display"],
                mode="lines",
                name=val_lbl,
                line=dict(color="#06d6f0", width=2.4),
                fill="tozeroy",
                fillcolor="rgba(6,214,240,0.22)",
                hovertemplate="Date: %{x|%d/%m/%Y}<br>Level: %{y:.3f}<extra></extra>",
            )
        )
    else:
        fig.add_trace(
            go.Scatter(
                x=dfp["date"],
                y=dfp["display"],
                mode="lines+markers",
                name=val_lbl,
                line=dict(color="#06d6f0", width=2.4),
                marker=dict(size=7, color="#0284c7", line=dict(width=1, color="white")),
                hovertemplate="Date: %{x|%d/%m/%Y}<br>Level: %{y:.3f}<extra></extra>",
            )
        )

    fig.update_layout(
        title=f"{val_lbl} στο χρόνο",
        xaxis=dict(title="Ημερομηνία", tickformat="%d/%m/%Y", tickangle=-30),
        yaxis=dict(title=val_lbl),
        height=380,
        plot_bgcolor=_PLT_BG,
        paper_bgcolor=_PLT_PAPER,
        font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
        margin=dict(t=60, b=30, l=60, r=20),
    )
    fig.update_xaxes(showgrid=True, gridcolor=_PLT_GRID, linecolor=_PLT_LINE)
    fig.update_yaxes(showgrid=True, gridcolor=_PLT_GRID, linecolor=_PLT_LINE)
    st.plotly_chart(fig, use_container_width=True, theme=None)

# ─── Top-level header and view selector ───────────────────────────────────────
render_main_header()

MAIN_VIEW_OPTIONS = ["Δορυφορικά δεδομένα", "Μετρήσεις πεδίου", "Παραδοτέα Έργου"]
try:
    selected_main_view = st.segmented_control(
        "Προβολή",
        options=MAIN_VIEW_OPTIONS,
        default=MAIN_VIEW_OPTIONS[0],
        selection_mode="single",
        key="main_view_selector",
    )
except Exception:
    selected_main_view = st.radio(
        "Προβολή",
        options=MAIN_VIEW_OPTIONS,
        index=0,
        horizontal=True,
        key="main_view_selector_fallback",
    )

if not selected_main_view:
    selected_main_view = MAIN_VIEW_OPTIONS[0]

if selected_main_view == "Δορυφορικά δεδομένα":
    render_satellite_data_view()
    st.stop()

if selected_main_view == "Παραδοτέα Έργου":
    render_deliverables_view()
    st.stop()

# ─── Load data ─────────────────────────────────────────────────────────────────
EXCEL_PATH = "ΑΠΟΤΕΛΕΣΜΑΤΑ_ΔΟΡΥΦΟΡΙΚΗΣ_ΠΑΡΑΚΟΛΟΥΘΗΣΗΣ_ΦΡΑΓΜΑΤΟΣ_2025-2026_ΕΥΑΘ.xlsx"

_use_light_plot_theme()

def _simplify_name(name):
    return re.sub(r"[\s_.-]+", "", name).lower()

def _unique_existing_dirs(paths):
    out = []
    seen = set()
    for p in paths:
        if p is None:
            continue
        pp = Path(p).expanduser()
        try:
            rp = pp.resolve()
        except Exception:
            rp = pp
        key = str(rp).lower()
        if key in seen:
            continue
        seen.add(key)
        if rp.exists() and rp.is_dir():
            out.append(rp)
    return out

def resolve_excel_path(preferred_name):
    preferred_path = Path(preferred_name).expanduser()
    if preferred_path.is_absolute() and preferred_path.exists():
        return preferred_path

    search_dirs = _unique_existing_dirs(
        [
            FIELD_DATA_ROOT,
            APP_DIR,
            PLATFORM_ROOT,
            PLATFORM_ROOT / "DATA",
        ]
    )

    for d in search_dirs:
        p = d / preferred_name
        if p.exists():
            return p

    target_key = _simplify_name(preferred_path.name)
    xlsx_files = []
    for d in search_dirs:
        xlsx_files.extend(sorted(d.rglob("*.xlsx")))

    dedup = {}
    for p in xlsx_files:
        try:
            dedup[str(p.resolve()).lower()] = p
        except Exception:
            dedup[str(p).lower()] = p
    xlsx_files = list(dedup.values())

    for candidate in xlsx_files:
        if _simplify_name(candidate.name) == target_key:
            return candidate

    year_matches = [path for path in xlsx_files if "2025" in path.name and "2026" in path.name]
    if len(year_matches) == 1:
        return year_matches[0]

    if len(xlsx_files) == 1:
        return xlsx_files[0]

    available = ", ".join(str(path) for path in xlsx_files) or "none"
    raise FileNotFoundError(
        f"Excel file '{preferred_name}' was not found. Search dirs: {search_dirs}. Available .xlsx files: {available}"
    )

EXCEL_PATH = resolve_excel_path(EXCEL_PATH)
MEASUREMENTS_SOURCE_PATH = Path(EXCEL_PATH).resolve()

try:
    df = load_data(EXCEL_PATH)
except PermissionError as e:
    st.error(
        f"Σφάλμα πρόσβασης στο αρχείο Excel: {e}. "
        "Κλείστε τυχόν ανοιχτό Excel στο ίδιο αρχείο και δοκιμάστε ξανά."
    )
    st.stop()
except Exception as e:
    st.error(f"Σφάλμα φόρτωσης: {e}")
    st.stop()

if df.empty or "date" not in df.columns:
    st.error("No valid measurements were parsed from the Excel workbook.")
    st.caption("Check the worksheet structure and point/depth rows, then rerun.")
    st.stop()

st.caption(f"Πηγή δεδομένων μετρήσεων πεδίου (Excel): `{MEASUREMENTS_SOURCE_PATH}`")

# ─── CSS ───────────────────────────────────────────────────────────────────────
# ── Dark theme CSS (matches streamlit_geotiff_map_1.py) ──────────────────────
CSS = """
<link href="https://fonts.googleapis.com/css2?family=Bricolage+Grotesque:opsz,wght@12..96,500;12..96,600;12..96,700;12..96,800&family=Plus+Jakarta+Sans:ital,wght@0,300;0,400;0,500;0,600&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0">
<style>
:root{
  --bg:#060d18;--bg2:#0a1525;--sf:#0e1e30;--sf2:#122236;
  --ac:#06d6f0;--acd:rgba(6,214,240,.12);--abdr:rgba(6,214,240,.22);
  --tx:#dff2fa;--mid:#6ab4ce;--dim:#2e6480;--bdr:rgba(6,214,240,.13);
  --sh:0 10px 52px rgba(0,0,0,.7);--r:16px;
  --fh:'Bricolage Grotesque',sans-serif;
  --fb:'Plus Jakarta Sans',sans-serif;
  --fm:'JetBrains Mono',monospace;
}
html,body,[data-testid="stApp"]{background:var(--bg)!important;color:var(--tx)!important;font-family:var(--fb)!important;}
[data-testid="stApp"]::before{content:'';position:fixed;inset:0;pointer-events:none;z-index:0;background-image:linear-gradient(rgba(6,214,240,.018) 1px,transparent 1px),linear-gradient(90deg,rgba(6,214,240,.018) 1px,transparent 1px);background-size:56px 56px;}
#MainMenu,footer,header,[data-testid="stDecoration"],[data-testid="stToolbar"]{display:none!important;}
.block-container{padding-top:1.4rem!important;padding-bottom:5rem!important;max-width:1480px!important;position:relative;z-index:1;}

.hcard{background:linear-gradient(140deg,#091726 0%,#0d2340 55%,#071520 100%);border:1px solid var(--abdr);border-top:2px solid rgba(6,214,240,.55);border-radius:var(--r);padding:1.6rem 2.5rem;margin-bottom:2rem;display:flex;align-items:center;gap:2.4rem;box-shadow:var(--sh),inset 0 1px 0 rgba(255,255,255,.04);position:relative;overflow:hidden;}
.hcard::before{content:'';position:absolute;top:-80px;right:-80px;width:260px;height:260px;background:radial-gradient(circle,rgba(6,214,240,.08) 0%,transparent 70%);pointer-events:none;}
.hcard h1{font-family:var(--fh)!important;font-size:1.45rem!important;font-weight:700!important;color:#f0faff!important;margin:0 0 .35rem 0!important;line-height:1.3!important;letter-spacing:-.02em!important;}
.hcard .sub{font-size:.72rem;color:var(--dim);letter-spacing:.1em;text-transform:uppercase;font-weight:500;}
.badge{display:inline-flex;align-items:center;gap:.4rem;background:var(--acd);border:1px solid var(--abdr);color:var(--ac);border-radius:99px;padding:.22rem .9rem;font-family:var(--fh);font-size:.65rem;font-weight:600;letter-spacing:.05em;margin-top:.5rem;}

.slabel{font-family:var(--fh);font-size:.6rem;font-weight:600;letter-spacing:.22em;text-transform:uppercase;color:var(--dim);margin-bottom:.75rem;padding-left:.15rem;}

.metric-card{background:linear-gradient(135deg,var(--sf),var(--bg2));border:1px solid var(--bdr);border-top:2px solid rgba(6,214,240,.32);border-radius:14px;padding:1.1rem 1.3rem;text-align:center;transition:border-color .2s,box-shadow .2s;}
.metric-card:hover{border-color:rgba(6,214,240,.42);box-shadow:0 0 28px rgba(6,214,240,.08);}
.metric-val{font-family:var(--fh);font-size:1.8rem;font-weight:700;color:var(--ac);letter-spacing:-.03em;}
.metric-label{font-family:var(--fh);font-size:.6rem;color:var(--dim);margin-top:.4rem;letter-spacing:.12em;text-transform:uppercase;font-weight:600;}

[data-testid="stMetric"]{background:linear-gradient(135deg,var(--sf),var(--bg2))!important;border:1px solid var(--bdr)!important;border-top:2px solid rgba(6,214,240,.3)!important;border-radius:14px!important;padding:1rem 1.2rem!important;}
[data-testid="stMetricLabel"]{font-family:var(--fh)!important;font-size:.6rem!important;letter-spacing:.12em!important;text-transform:uppercase!important;color:var(--dim)!important;font-weight:600!important;}
[data-testid="stMetricValue"]{font-family:var(--fh)!important;font-size:1.6rem!important;color:var(--ac)!important;font-weight:700!important;letter-spacing:-.02em!important;}

[data-testid="stButton"]>button{background:var(--sf)!important;border:1px solid var(--bdr)!important;color:var(--mid)!important;border-radius:12px!important;font-family:var(--fh)!important;font-size:.75rem!important;font-weight:600!important;padding:.7rem 1rem!important;transition:all .18s ease!important;}
[data-testid="stButton"]>button:hover{background:var(--sf2)!important;border-color:rgba(6,214,240,.5)!important;color:var(--tx)!important;transform:translateY(-1px)!important;box-shadow:0 6px 22px rgba(0,0,0,.4)!important;}
[data-testid="stDownloadButton"]>button{background:linear-gradient(135deg,#073d60,#052e4a)!important;border-color:var(--ac)!important;color:#e0f8ff!important;}

[data-testid="stSelectbox"] label,[data-testid="stMultiselect"] label,[data-testid="stSlider"] label,[data-testid="stDateInput"] label,[data-testid="stRadio"] label{font-family:var(--fh)!important;font-size:.62rem!important;letter-spacing:.13em!important;text-transform:uppercase!important;color:var(--dim)!important;font-weight:600!important;}
[data-testid="stSelectbox"] [data-baseweb="select"]>div{background:var(--bg2)!important;border-color:var(--bdr)!important;color:var(--mid)!important;border-radius:8px!important;}
[data-baseweb="slider"] [role="slider"]{background:var(--ac)!important;box-shadow:0 0 0 3px rgba(6,214,240,.2)!important;}

[data-testid="stTabs"] [role="tablist"]{border-bottom:1px solid var(--bdr)!important;gap:.2rem!important;}
[data-testid="stTabs"] [role="tab"]{font-family:var(--fh)!important;font-size:.7rem!important;font-weight:600!important;letter-spacing:.07em!important;text-transform:uppercase!important;color:var(--dim)!important;padding:.5rem 1.1rem!important;border-radius:8px 8px 0 0!important;transition:all .15s!important;}
[data-testid="stTabs"] [role="tab"]:hover{color:var(--mid)!important;}
[data-testid="stTabs"] [role="tab"][aria-selected="true"]{color:var(--ac)!important;border-bottom:2px solid var(--ac)!important;background:var(--acd)!important;}

[data-testid="stExpander"]{background:var(--sf)!important;border:1px solid var(--bdr)!important;border-radius:12px!important;}
[data-testid="stExpander"] summary{font-family:var(--fh)!important;font-size:.8rem!important;font-weight:600!important;color:var(--mid)!important;}

[data-testid="stDataFrame"] thead th{background:var(--sf2)!important;color:var(--mid)!important;font-family:var(--fh)!important;font-size:.65rem!important;letter-spacing:.08em!important;text-transform:uppercase!important;}

[data-testid="stSidebar"]{background:var(--bg2)!important;border-right:1px solid var(--bdr)!important;}
[data-testid="stSidebar"] h3{font-family:var(--fh)!important;font-size:.72rem!important;font-weight:700!important;letter-spacing:.15em!important;text-transform:uppercase!important;color:var(--dim)!important;}

[data-testid="stCaptionContainer"]{color:var(--dim)!important;font-size:.68rem!important;font-family:var(--fm)!important;}
[data-testid="stInfo"]{background:rgba(6,214,240,.06)!important;border:1px solid rgba(6,214,240,.2)!important;border-radius:10px!important;color:var(--mid)!important;}
[data-testid="stWarning"]{background:rgba(245,158,11,.07)!important;border:1px solid rgba(245,158,11,.25)!important;border-radius:10px!important;}
[data-testid="stError"]{background:rgba(239,68,68,.07)!important;border:1px solid rgba(239,68,68,.25)!important;border-radius:10px!important;}

::-webkit-scrollbar{width:5px;height:5px;}
::-webkit-scrollbar-track{background:var(--bg);}
::-webkit-scrollbar-thumb{background:#1a3d58;border-radius:3px;}
::-webkit-scrollbar-thumb:hover{background:#2a5472;}
hr{border-color:var(--bdr)!important;}

/* ═══════════════════════════════════════════════════════════════
   MOBILE RESPONSIVE — screens < 768px
   ═══════════════════════════════════════════════════════════════ */
@media (max-width: 768px) {

  /* ── Layout ──────────────────────────────────────────────── */
  .block-container {
    padding-top: .8rem !important;
    padding-left: .75rem !important;
    padding-right: .75rem !important;
    padding-bottom: 5rem !important;
  }

  /* ── Header: stack logo above text on small screens ──────── */
  .hcard {
    flex-direction: column !important;
    align-items: flex-start !important;
    gap: 1rem !important;
    padding: 1.2rem 1.4rem !important;
    border-radius: 14px !important;
    margin-bottom: 1.2rem !important;
  }
  .hcard img { height: 44px !important; }
  .hcard h1  { font-size: 1.1rem !important; }
  .hcard .sub{ font-size: .65rem !important; }

  /* ── Streamlit columns → stack vertically ────────────────── */
  [data-testid="stHorizontalBlock"] {
    flex-direction: column !important;
    gap: .75rem !important;
  }
  [data-testid="stColumn"] {
    width: 100% !important;
    min-width: 100% !important;
    flex: 1 1 100% !important;
  }

  /* ── Metric cards ─────────────────────────────────────────── */
  .metric-card {
    padding: .9rem 1rem !important;
    border-radius: 12px !important;
  }
  .metric-val { font-size: 1.5rem !important; }
  .metric-label { font-size: .58rem !important; }

  /* ── Buttons — bigger touch targets ──────────────────────── */
  [data-testid="stButton"] > button {
    min-height: 48px !important;
    font-size: .78rem !important;
    padding: .8rem 1rem !important;
    border-radius: 10px !important;
    width: 100% !important;
  }

  /* ── Tabs — scrollable on mobile ─────────────────────────── */
  [data-testid="stTabs"] [role="tablist"] {
    overflow-x: auto !important;
    flex-wrap: nowrap !important;
    -webkit-overflow-scrolling: touch !important;
    scrollbar-width: none !important;
    padding-bottom: 2px !important;
  }
  [data-testid="stTabs"] [role="tablist"]::-webkit-scrollbar { display:none; }
  [data-testid="stTabs"] [role="tab"] {
    white-space: nowrap !important;
    font-size: .65rem !important;
    padding: .45rem .85rem !important;
    min-width: fit-content !important;
  }

  /* ── Select boxes & inputs ────────────────────────────────── */
  [data-testid="stSelectbox"] [data-baseweb="select"] > div,
  [data-testid="stDateInput"] input {
    font-size: .85rem !important;
    min-height: 44px !important;
  }

  /* ── Map — full width, reasonable height ──────────────────── */
  .mapwrap { border-radius: 12px !important; }
  [data-testid="stIFrame"] { min-height: 340px !important; }
  iframe { min-height: 340px !important; }

  /* ── Status strip ─────────────────────────────────────────── */
  .sstrip {
    font-size: .72rem !important;
    padding: .4rem .75rem !important;
  }

  /* ── Section labels ───────────────────────────────────────── */
  .slabel { font-size: .58rem !important; letter-spacing: .16em !important; }

  /* ── Sidebar: auto-hides on mobile in Streamlit ──────────── */
  [data-testid="stSidebar"] > div:first-child {
    padding-top: 1rem !important;
  }

  /* ── Plotly charts ────────────────────────────────────────── */
  [data-testid="stPlotlyChart"] { overflow-x: auto !important; }
  [data-testid="stPlotlyChart"] > div { min-width: 0 !important; }

  /* ── Caption/mono text ────────────────────────────────────── */
  [data-testid="stCaptionContainer"] { font-size: .62rem !important; }

  /* ── pydeck/folium map containers ────────────────────────── */
  [data-testid="stDeckGlJsonChart"],
  .stFolium { border-radius: 12px !important; overflow: hidden !important; }

  /* ── Expander ─────────────────────────────────────────────── */
  [data-testid="stExpander"] summary {
    font-size: .75rem !important;
    padding: .8rem !important;
  }
}

/* ── Extra-small phones (< 420px) ──────────────────────────────── */
@media (max-width: 420px) {
  .hcard h1  { font-size: .95rem !important; }
  .metric-val { font-size: 1.35rem !important; }
  [data-testid="stTabs"] [role="tab"] {
    font-size: .6rem !important;
    padding: .4rem .7rem !important;
  }
}
</style>
"""
st.markdown(CSS, unsafe_allow_html=True)

# ── Mobile detection via JS ────────────────────────────────────────────────────
_MOBILE_JS = """
<script>
(function() {
    const w = window.innerWidth;
    const mobile = w < 768;
    // Send to Streamlit via URL param trick — use sessionStorage flag
    if (mobile !== (sessionStorage.getItem('_st_mobile') === '1')) {
        sessionStorage.setItem('_st_mobile', mobile ? '1' : '0');
        // Set a cookie Streamlit can read via query params on reload
        const url = new URL(window.location.href);
        url.searchParams.set('_mobile', mobile ? '1' : '0');
        window.location.replace(url.toString());
    }
})();
</script>
"""
st.markdown(_MOBILE_JS, unsafe_allow_html=True)

# Read mobile flag from query params (set by JS above)
_qp = st.query_params
_is_mobile = str(_qp.get("_mobile", "0")) == "1"

def _cols(*desktop_spec):
    """Return st.columns with mobile-aware spec.
    On mobile, always returns a single column."""
    if _is_mobile:
        return st.columns(1)
    return st.columns(list(desktop_spec) if len(desktop_spec) > 1 else desktop_spec[0])



st.markdown("---")

# ─── Summary metrics ───────────────────────────────────────────────────────────
col1, col2, col3, col4 = (st.columns(2) + st.columns(2)) if _is_mobile else st.columns(4)
dates = sorted(df["date"].unique())
with col1:
    st.markdown(f"""<div class="metric-card"><div style="font-size:1.3rem;margin-bottom:.3rem">📅</div><div class="metric-val">{len(dates)}</div><div class="metric-label">Δειγματοληψίες</div></div>""", unsafe_allow_html=True)
with col2:
    st.markdown(f"""<div class="metric-card"><div style="font-size:1.3rem;margin-bottom:.3rem">📍</div><div class="metric-val">{df['point'].nunique()}</div><div class="metric-label">Σημεία</div></div>""", unsafe_allow_html=True)
with col3:
    st.markdown(f"""<div class="metric-card"><div style="font-size:1.3rem;margin-bottom:.3rem">🌊</div><div class="metric-val">{df['depth'].nunique()}</div><div class="metric-label">Βάθη</div></div>""", unsafe_allow_html=True)
with col4:
    params_with_data = sum(1 for p in COL_MAP if df[p].notna().any())
    st.markdown(f"""<div class="metric-card"><div style="font-size:1.3rem;margin-bottom:.3rem">🔬</div><div class="metric-val">{params_with_data}</div><div class="metric-label">Παράμετροι</div></div>""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# Shared map data + figure builder
map_df = pd.DataFrame([
    {
        "Σημείο": f"Σ{k}",
        "id": k,
        "lat": v["lat"],
        "lon": v["lon"],
        "Μετρήσεις": int(df[df["point"] == k].shape[0]),
    }
    for k, v in SAMPLING_POINTS.items()
])

# Dynamic parameter-based map (overrides the simple counts-only version above)
MAP_POINTS_DF = pd.DataFrame([
    {"point_id": k, "point_label": f"Σ{k}", "lat": v["lat"], "lon": v["lon"]}
    for k, v in SAMPLING_POINTS.items()
])

def _build_map_dataset(color_param="__counts__", sel_date=None, sel_depth="all"):
    map_data = MAP_POINTS_DF.copy()
    point_counts = df.groupby("point").size()
    map_data["measurement_count"] = map_data["point_id"].map(point_counts).fillna(0).astype(int)
    map_data["param_value"] = np.nan

    if color_param != "__counts__":
        sub = df.copy()
        if sel_date is not None:
            sub = sub[sub["date"] == sel_date]
        if sel_depth != "all":
            sub = sub[sub["depth"] == sel_depth]

        agg = (
            sub.groupby("point")[color_param]
            .mean()
            .reset_index()
            .rename(columns={"point": "point_id", color_param: "param_value"})
        )
        map_data = map_data.merge(agg, on="point_id", how="left", suffixes=("", "_new"))
        if "param_value_new" in map_data.columns:
            map_data["param_value"] = map_data["param_value_new"]
            map_data = map_data.drop(columns=["param_value_new"])

    return map_data

def build_sampling_map(
    height=550,
    title=None,
    color_param="__counts__",
    sel_date=None,
    sel_depth="all",
    show_colorbar=True,
    zoom_override=None,
):
    map_data = _build_map_dataset(color_param=color_param, sel_date=sel_date, sel_depth=sel_depth)
    use_counts = color_param == "__counts__" or map_data["param_value"].notna().sum() == 0
    color_col = "measurement_count" if use_counts else "param_value"
    color_label = "Μετρήσεις" if use_counts else color_param
    range_color = None

    if use_counts:
        cmin = float(map_data["measurement_count"].min())
        cmax = float(map_data["measurement_count"].max())
        if cmax > cmin:
            range_color = (cmin, cmax)
    else:
        all_vals = pd.to_numeric(df[color_param], errors="coerce")
        vmin = all_vals.min(skipna=True)
        vmax = all_vals.max(skipna=True)
        if pd.notna(vmin) and pd.notna(vmax):
            vmin = float(vmin)
            vmax = float(vmax)
            if vmax <= vmin:
                vmax = vmin + 1e-9
            range_color = (vmin, vmax)

    lat_min, lat_max = map_data["lat"].min(), map_data["lat"].max()
    lon_min, lon_max = map_data["lon"].min(), map_data["lon"].max()
    map_center = {"lat": float((lat_min + lat_max) / 2), "lon": float((lon_min + lon_max) / 2)}
    span = max(float(lat_max - lat_min), float(lon_max - lon_min))
    if span > 0.12:
        zoom_level = 10.2
    elif span > 0.06:
        zoom_level = 10.9
    elif span > 0.03:
        zoom_level = 11.4
    else:
        zoom_level = 12.0
    if height <= 320:
        zoom_level -= 0.9
    if zoom_override is not None:
        zoom_level = zoom_override

    hover = {
        "point_label": False,
        "lat": ":.6f",
        "lon": ":.6f",
        "measurement_count": True,
    }
    if not use_counts:
        hover["param_value"] = ":.3f"

    fig_map = px.scatter_mapbox(
        map_data,
        lat="lat",
        lon="lon",
        text="point_label",
        color=color_col,
        range_color=range_color,
        size=[20] * len(map_data),
        color_continuous_scale="Turbo" if not use_counts else "Blues",
        hover_name="point_label",
        hover_data=hover,
        zoom=zoom_level,
        center=map_center,
        mapbox_style="open-street-map",
        title=title,
        height=height,
        labels={
            "point_label": "Σημείο",
            "measurement_count": "Μετρήσεις",
            "param_value": color_label,
            "lat": "Γεωγρ. Πλάτος",
            "lon": "Γεωγρ. Μήκος",
        },
    )
    fig_map.update_traces(marker=dict(size=16), textposition="top center", textfont=dict(size=13, color="black"))
    if show_colorbar:
        fig_map.update_coloraxes(
            colorbar_title=color_label,
            colorbar=dict(thickness=14, len=0.75),
        )
    else:
        fig_map.update_layout(coloraxis_showscale=False)
    fig_map.update_layout(margin={"r": 0, "t": 40 if title else 10, "l": 0, "b": 0})
    return fig_map

def build_3d_lake_deck(
    data_frame,
    param_name,
    global_min=None,
    global_max=None,
    pitch=58,
    bearing=28,
    zoom_override=None,
    depth_exaggeration=1.2,
    map_theme="Voyager",
    show_labels=True,
    drag_mode="rotate",
):
    if data_frame.empty:
        return None, np.nan, np.nan

    deck_df = data_frame.copy()
    deck_df[param_name] = pd.to_numeric(deck_df[param_name], errors="coerce")
    deck_df["depth_m"] = deck_df["depth"].apply(depth_to_m)
    deck_df = deck_df.dropna(subset=["date", "point", "depth_m", param_name])
    if deck_df.empty:
        return None, np.nan, np.nan

    deck_df["point"] = deck_df["point"].astype(int)
    deck_df["point_label"] = deck_df["point"].apply(lambda p: f"Σ{p}")
    deck_df["date_label"] = pd.to_datetime(deck_df["date"]).dt.strftime("%d/%m/%Y")
    deck_df["date"] = pd.to_datetime(deck_df["date"]).dt.strftime("%Y-%m-%d")
    deck_df = deck_df.merge(
        MAP_POINTS_DF[["point_id", "lat", "lon"]],
        left_on="point",
        right_on="point_id",
        how="left",
    )
    deck_df = deck_df.dropna(subset=["lat", "lon"])
    if deck_df.empty:
        return None, np.nan, np.nan

    if global_min is None or global_max is None:
        pvals = pd.to_numeric(df[param_name], errors="coerce")
        pmin = float(pvals.min(skipna=True))
        pmax = float(pvals.max(skipna=True))
    else:
        pmin = float(global_min)
        pmax = float(global_max)
    if pmax <= pmin:
        pmax = pmin + 1e-9

    is_multi_date = deck_df["date"].nunique() > 1
    marker_alpha = 185 if is_multi_date else 228
    deck_df["color_rgba"] = _values_to_rgba(deck_df[param_name], pmin, pmax, colorscale="Turbo", alpha=marker_alpha)

    depth_scale = 16.0 * float(depth_exaggeration)
    max_depth = float(deck_df["depth_m"].max()) if not deck_df["depth_m"].isna().all() else 1.0
    z_surface = (max_depth + 1.0) * depth_scale
    deck_df["z_pos"] = z_surface - deck_df["depth_m"].clip(lower=0) * depth_scale
    deck_df["param_value"] = deck_df[param_name].round(3)
    deck_df["depth_label"] = deck_df["depth_m"].round(2)
    deck_df["z_label"] = (-deck_df["depth_m"]).round(2)

    # Reduce overplotting in all-dates mode with a tiny circular jitter per date.
    if is_multi_date:
        code = pd.Categorical(pd.to_datetime(deck_df["date"]).dt.strftime("%Y-%m-%d")).codes.astype(float)
        jitter = 0.00010
        angle = code * 0.9
        deck_df["lon_plot"] = deck_df["lon"] + np.cos(angle) * jitter
        deck_df["lat_plot"] = deck_df["lat"] + np.sin(angle) * jitter
    else:
        deck_df["lon_plot"] = deck_df["lon"]
        deck_df["lat_plot"] = deck_df["lat"]

    station_df = (
        deck_df.groupby(["point", "point_label", "lat", "lon"], as_index=False)["depth_m"]
        .max()
        .rename(columns={"depth_m": "max_depth_m"})
    )
    station_df["src"] = station_df.apply(
        lambda r: [float(r["lon"]), float(r["lat"]), float(z_surface)],
        axis=1
    )
    station_df["dst"] = station_df.apply(
        lambda r: [float(r["lon"]), float(r["lat"]), float(z_surface - float(r["max_depth_m"]) * depth_scale)],
        axis=1
    )
    station_df["z_surface"] = z_surface
    station_df["z_text"] = z_surface + 18.0

    lat_span = float(MAP_POINTS_DF["lat"].max() - MAP_POINTS_DF["lat"].min()) * 2.3
    lon_span = float(MAP_POINTS_DF["lon"].max() - MAP_POINTS_DF["lon"].min()) * 2.3
    span = max(lat_span, lon_span)
    if zoom_override is not None:
        zoom = float(zoom_override)
    elif span > 0.16:
        zoom = 10.4
    elif span > 0.10:
        zoom = 10.8
    elif span > 0.06:
        zoom = 11.1
    else:
        zoom = 11.4

    view_state = pdk.ViewState(
        latitude=float(MAP_POINTS_DF["lat"].mean()),
        longitude=float(MAP_POINTS_DF["lon"].mean()),
        zoom=zoom,
        pitch=float(pitch),
        bearing=float(bearing),
    )

    layers = [
        pdk.Layer(
            "LineLayer",
            data=station_df.to_dict(orient="records"),
            get_source_position="src",
            get_target_position="dst",
            get_width=4,
            width_min_pixels=2,
            get_color=[100, 116, 139, 210],
            pickable=False,
        ),
        pdk.Layer(
            "ScatterplotLayer",
            data=deck_df.to_dict(orient="records"),
            get_position="[lon_plot, lat_plot, z_pos]",
            get_radius=95,
            radius_min_pixels=4,
            get_fill_color="color_rgba",
            stroked=False,
            pickable=True,
            auto_highlight=True,
        ),
    ]

    if show_labels:
        layers.append(
            pdk.Layer(
                "TextLayer",
                data=station_df.to_dict(orient="records"),
                get_position="[lon, lat, z_text]",
                get_text="point_label",
                get_size=15,
                get_color=[22, 28, 45, 255],
                get_angle=0,
                get_text_anchor="'middle'",
                get_alignment_baseline="'bottom'",
                pickable=False,
            )
        )

    map_style_lookup = {
        "Voyager": "https://basemaps.cartocdn.com/gl/voyager-gl-style/style.json",
        "Positron": "https://basemaps.cartocdn.com/gl/positron-gl-style/style.json",
        "Dark Matter": "https://basemaps.cartocdn.com/gl/dark-matter-gl-style/style.json",
    }
    map_style = map_style_lookup.get(map_theme, map_style_lookup["Voyager"])

    deck = pdk.Deck(
        layers=layers,
        initial_view_state=view_state,
        views=[pdk.View(
            type="MapView",
            controller={
                "dragRotate": True,
                "touchRotate": True,
                "dragPan": True,
                "scrollZoom": True,
                "dragMode": drag_mode,
            }
        )],
        map_style=map_style,
        tooltip={
            "html": (
                "<b>{point_label}</b><br/>"
                "Ημερομηνία: {date_label}<br/>"
                "Βάθος: {depth_label} m<br/>"
                "z: {z_label} m<br/>"
                f"{param_name}: {{param_value}}"
            ),
            "style": {"backgroundColor": "#0f172a", "color": "white"},
        },
    )
    return deck, pmin, pmax

# Keep the map visible alongside all tabs/charts
map_param_options = [p for p in COL_MAP if df[p].notna().any()] + ["Μετρήσεις"]
map_date_options = ["Όλες"] + [d.strftime("%d/%m/%Y") for d in dates]
map_depth_options = ["Όλα"] + [d for d in DEPTH_ORDER if d in df["depth"].unique()]

with st.sidebar:
    map_param_choice = st.selectbox("Χρωματισμός Παραμέτρου", options=map_param_options, key="map_color_param")
    map_date_choice = st.selectbox("Ημερομηνία", options=map_date_options, key="map_color_date")
    map_depth_choice = st.selectbox("Βάθος", options=map_depth_options, key="map_color_depth")

map_param_key = "__counts__" if map_param_choice == "Μετρήσεις" else map_param_choice
map_date_value = None if map_date_choice == "Όλες" else pd.to_datetime(map_date_choice, format="%d/%m/%Y")
map_depth_value = "all" if map_depth_choice == "Όλα" else map_depth_choice

with st.sidebar:
    st.markdown("### 🗺️ Χάρτης Σημείων")
    st.plotly_chart(
        build_sampling_map(
            height=280,
            title=None,
            color_param=map_param_key,
            sel_date=map_date_value,
            sel_depth=map_depth_value,
            show_colorbar=True,
            zoom_override=10.8,
        ),
        use_container_width=True,
        theme=None
    )
    legend_df = _build_map_dataset(
        color_param=map_param_key,
        sel_date=map_date_value,
        sel_depth=map_depth_value,
    ).sort_values("point_id")
    if map_param_key == "__counts__":
        legend_view = legend_df[["point_label", "measurement_count"]].rename(columns={
            "point_label": "Σημείο",
            "measurement_count": "Μετρήσεις",
        })
    else:
        legend_df[map_param_choice] = legend_df["param_value"].round(3)
        legend_df[map_param_choice] = legend_df[map_param_choice].where(
            legend_df[map_param_choice].notna(),
            "-"
        )
        legend_view = legend_df[["point_label", map_param_choice]].rename(columns={
            "point_label": "Σημείο",
        })
    st.markdown("**Legend**")
    st.dataframe(legend_view, use_container_width=True, hide_index=True, height=220)
    st.caption("Ο χάρτης παραμένει ορατός σε όλα τα διαγράμματα.")

# ─── TABS ──────────────────────────────────────────────────────────────────────
GEE_PARAMETER_OPTIONS = [
    "Water Surface Temperature - Surface Temperature (°C)",
    "True Color - RGB",
    "NDCI - Chlorophyll Proxy",
    "NDTI - Turbidity Proxy",
    "NDWI - Water Index",
    "GNIR - Green/NIR Ratio",
    "CDOM - Colored Dissolved Organic Matter",
    "TSM - Total Suspended Matter",
    "Chl-a - Chlorophyll-a (estimate)",
    "FAI - Floating Algae Index",
    "Algae Bloom Detection",
    "Water Turbidity Classes",
    "Anomaly Detection",
    "Chl-a (Se2WaQ)",
    "Cya (Se2WaQ)",
    "Turb (Se2WaQ)",
    "CDOM (Se2WaQ)",
    "DOC (Se2WaQ)",
    "Color (Se2WaQ)",
    "Chl-a (Gadouras Poly2)",
    "Chl-a (Gadouras Linear)",
    "Chl-a (Gadouras LogLin)",
    "Chl-a (Gadouras Autumn)",
    "NDCI (Gadouras scale)",
]

GEE_LAKE_METHODS = [
    "Auto (Best)",
    "NDWI Standard",
    "MNDWI Enhanced",
    "Multi-Index Fusion",
    "AWEInsh",
    "Simple Threshold",
]

DEFAULT_EE_PROJECT = os.getenv("EE_DEFAULT_PROJECT", "ee-lioumbas")


def _discover_local_service_account() -> Optional[Dict[str, str]]:
    base_dir = Path(__file__).resolve().parent
    candidate_paths = list(base_dir.glob("ee-*.json")) + list(base_dir.glob("*service*.json"))
    for path in sorted(candidate_paths):
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if payload.get("type") != "service_account":
            continue
        return {
            "path": str(path),
            "email": str(payload.get("client_email", "")),
            "project_id": str(payload.get("project_id", "")),
            "key_json": json.dumps(payload),
        }
    return None


def _init_ee(
    project_id: Optional[str] = None,
    svc_account_email: Optional[str] = None,
    svc_key_json: Optional[str] = None,
    allow_auto_service_account: bool = False,
) -> Tuple[bool, str]:
    if ee is None:
        return False, "earthengine-api is not installed in this environment."

    diagnostics = []
    discovered_sa = _discover_local_service_account() if allow_auto_service_account else None
    if discovered_sa:
        diagnostics.append(f"Detected local service-account key: {discovered_sa.get('path')}")
        if not svc_key_json:
            svc_key_json = discovered_sa.get("key_json")
        if not svc_account_email:
            svc_account_email = discovered_sa.get("email")

    # Env fallbacks
    if not svc_key_json:
        for env_key in ("EE_SERVICE_ACCOUNT_JSON", "GOOGLE_CREDENTIALS_JSON"):
            env_val = os.getenv(env_key, "").strip()
            if env_val:
                svc_key_json = env_val
                diagnostics.append(f"Using service-account JSON from env: {env_key}")
                break
    svc_key_path = ""
    if not svc_key_json:
        for env_key in ("EE_DEFAULT_KEY_PATH", "GOOGLE_APPLICATION_CREDENTIALS"):
            env_val = os.getenv(env_key, "").strip()
            if env_val:
                expanded = os.path.expanduser(env_val)
                if os.path.isfile(expanded):
                    svc_key_path = expanded
                    diagnostics.append(f"Using service-account key file from env: {env_key}={expanded}")
                    break
    if not svc_account_email:
        svc_account_email = os.getenv("EE_SERVICE_ACCOUNT", "").strip() or None

    creds = None
    detected_project = ""
    if svc_key_json:
        try:
            parsed = json.loads(svc_key_json)
            if not svc_account_email:
                svc_account_email = str(parsed.get("client_email", "")).strip() or None
            detected_project = str(parsed.get("project_id", "")).strip()
            if svc_account_email:
                creds = ee.ServiceAccountCredentials(svc_account_email, key_data=svc_key_json)
        except Exception as exc:
            diagnostics.append(f"Service-account JSON parse/auth failed: {exc}")
    elif svc_key_path and svc_account_email:
        try:
            creds = ee.ServiceAccountCredentials(svc_account_email, key_file=svc_key_path)
        except Exception as exc:
            diagnostics.append(f"Service-account file auth failed: {exc}")

    project_candidates = []
    for cand in [
        (project_id or "").strip(),
        os.getenv("EE_PROJECT", "").strip(),
        os.getenv("EARTHENGINE_PROJECT", "").strip(),
        (detected_project or "").strip(),
        (discovered_sa or {}).get("project_id", "").strip() if discovered_sa else "",
        DEFAULT_EE_PROJECT.strip(),
    ]:
        if cand and cand not in project_candidates:
            project_candidates.append(cand)

    attempts = []
    # Prefer explicit/known projects, then fallback to default/no-project.
    for project in project_candidates + [None]:
        try:
            kwargs = {}
            if creds is not None:
                kwargs["credentials"] = creds
            if project:
                kwargs["project"] = project
            ee.Initialize(**kwargs)
            try:
                ee.data.setDeadline(120000)
            except Exception:
                pass
            msg = ""
            if diagnostics:
                msg = " | ".join(diagnostics)
            if project:
                msg = (msg + " | " if msg else "") + f"Initialized with project: {project}"
            else:
                msg = (msg + " | " if msg else "") + "Initialized without explicit project"
            return True, msg
        except Exception as exc:
            ptxt = project if project else "<none>"
            attempts.append(f"project={ptxt}: {exc}")

    detail = "\n".join(attempts[:4])
    if diagnostics:
        detail = ("; ".join(diagnostics) + ("\n" + detail if detail else ""))
    return False, detail or "Earth Engine initialization failed."


def _mask_l8_sr(image):
    cloud_shadow = 1 << 3
    clouds = 1 << 5
    qa = image.select("QA_PIXEL")
    mask = qa.bitwiseAnd(cloud_shadow).eq(0).And(qa.bitwiseAnd(clouds).eq(0))
    return image.updateMask(mask)


def _safe_nd(image, band1: str, band2: str):
    return image.select(band1).subtract(image.select(band2)).divide(
        image.select(band1).add(image.select(band2)).add(ee.Image.constant(1e-4))
    )


def _apply_mask(image, mask, apply_mask: bool = True):
    return image.updateMask(mask) if apply_mask else image


def _process_gee_image(image, parameter: str, apply_mask: bool = True):
    if "Temperature" in parameter:
        water_mask = image.select("QA_PIXEL").bitwiseAnd(1 << 7).gt(0)
    else:
        water_mask = _safe_nd(image, "B3", "B8").gt(0)

    if parameter.startswith("Water Surface Temperature"):
        result = image.select("ST_B10").multiply(0.00341802).add(149.0).subtract(273.15).rename("WST")
    elif parameter.startswith("True Color"):
        return image.updateMask(water_mask)
    elif parameter.startswith("NDWI"):
        result = _safe_nd(image, "B3", "B8").rename("NDWI")
    elif parameter.startswith("NDCI (Gadouras scale)"):
        result = _safe_nd(image, "B5", "B4").rename("NDCI_Gadouras")
    elif parameter.startswith("NDCI"):
        result = _safe_nd(image, "B5", "B4").rename("NDCI")
    elif parameter.startswith("NDTI"):
        result = _safe_nd(image, "B4", "B3").rename("NDTI")
    elif parameter.startswith("GNIR"):
        result = image.select("B3").divide(image.select("B8").add(1e-4)).rename("GNIR")
    elif parameter.startswith("CDOM (Se2WaQ)"):
        result = ee.Image(537).multiply(
            ee.Image(-2.93).multiply(image.select("B3").divide(image.select("B4").add(1e-4))).exp()
        ).rename("CDOM_Se2WaQ")
    elif parameter.startswith("CDOM"):
        result = image.select("B3").divide(image.select("B2").add(1e-4)).rename("CDOM")
    elif parameter.startswith("TSM"):
        result = image.select("B4").multiply(0.0001).multiply(745.89).add(10.15).rename("TSM")
    elif parameter.startswith("Chl-a (Se2WaQ)"):
        result = ee.Image(4.26).multiply(
            image.select("B3").divide(image.select("B1").add(1e-4)).pow(3.94)
        ).rename("Chla_Se2WaQ")
    elif parameter.startswith("Cya (Se2WaQ)"):
        result = ee.Image(115530.31).multiply(
            image.select("B3").multiply(image.select("B4")).divide(image.select("B2").add(1e-4)).pow(2.38)
        ).rename("Cya_Se2WaQ")
    elif parameter.startswith("Turb (Se2WaQ)"):
        result = ee.Image(8.93).multiply(image.select("B3").divide(image.select("B1").add(1e-4))).subtract(6.39).rename("Turb_Se2WaQ")
    elif parameter.startswith("DOC (Se2WaQ)"):
        result = ee.Image(432).multiply(
            ee.Image(-2.24).multiply(image.select("B3").divide(image.select("B4").add(1e-4))).exp()
        ).rename("DOC_Se2WaQ")
    elif parameter.startswith("Color (Se2WaQ)"):
        result = ee.Image(25366).multiply(
            ee.Image(-4.53).multiply(image.select("B3").divide(image.select("B4").add(1e-4))).exp()
        ).rename("Color_Se2WaQ")
    elif parameter.startswith("Chl-a (Gadouras Poly2)"):
        ndci = _safe_nd(image, "B5", "B4")
        result = ndci.multiply(-61.626324).add(ee.Image.constant(10.348804)).add(ndci.pow(2).multiply(-58.284377)).rename("Chla_Poly2")
    elif parameter.startswith("Chl-a (Gadouras Linear)"):
        ndci = _safe_nd(image, "B5", "B4")
        result = ndci.multiply(-3.859564).add(ee.Image.constant(13.522046)).rename("Chla_Linear")
    elif parameter.startswith("Chl-a (Gadouras LogLin)"):
        ndci = _safe_nd(image, "B5", "B4")
        result = ndci.multiply(-0.369107).add(ee.Image.constant(2.462013)).exp().rename("Chla_LogLin")
    elif parameter.startswith("Chl-a (Gadouras Autumn)"):
        ndci = _safe_nd(image, "B5", "B4")
        result = ndci.multiply(-19.866480).add(ee.Image.constant(19.829734)).rename("Chla_Autumn")
    elif parameter.startswith("Chl-a -"):
        result = image.select("B5").divide(image.select("B4").add(1e-4)).pow(3.94).multiply(4.26).rename("Chl-a")
    elif parameter.startswith("FAI"):
        red = image.select("B4")
        nir = image.select("B8")
        swir = image.select("B11")
        baseline = red.add(swir.subtract(red).multiply((842 - 665) / (1610 - 665)))
        result = nir.subtract(baseline).rename("FAI")
    elif parameter.startswith("Algae Bloom Detection"):
        ndci = _safe_nd(image, "B5", "B4")
        chl = image.select("B5").divide(image.select("B4").add(1e-4)).pow(3.94).multiply(4.26)
        red = image.select("B4")
        nir = image.select("B8")
        swir = image.select("B11")
        baseline = red.add(swir.subtract(red).multiply((842 - 665) / (1610 - 665)))
        fai = nir.subtract(baseline)
        result = (
            ee.Image(0)
            .where(ndci.gt(0.1).And(chl.gt(5)), 1)
            .where(ndci.gt(0.3).And(chl.gt(15)), 2)
            .where(ndci.gt(0.5).And(chl.gt(25)).Or(fai.gt(0.02)), 3)
            .rename("AlgaeBloom")
        )
    elif parameter.startswith("Water Turbidity Classes"):
        ndti = _safe_nd(image, "B4", "B3")
        tsm = image.select("B4").multiply(0.0001).multiply(745.89).add(10.15)
        result = (
            ee.Image(1)
            .where(ndti.gt(-0.1).And(tsm.gt(10)), 2)
            .where(ndti.gt(0).And(tsm.gt(25)), 3)
            .where(ndti.gt(0.1).And(tsm.gt(50)), 4)
            .where(ndti.gt(0.2).And(tsm.gt(75)), 5)
            .rename("TurbidityClass")
        )
    elif parameter.startswith("Anomaly Detection"):
        ndci = _safe_nd(image, "B5", "B4")
        ndti = _safe_nd(image, "B4", "B3")
        cdom = image.select("B3").divide(image.select("B2").add(1e-4))
        result = (
            ee.Image(0)
            .where(ndci.gt(0.7).Or(ndci.lt(-0.3)), 1)
            .where(ndti.gt(0.4).Or(ndti.lt(-0.3)), 1)
            .where(cdom.gt(8).Or(cdom.lt(0.5)), 1)
            .rename("Anomaly")
        )
    else:
        result = _safe_nd(image, "B5", "B4").rename("NDCI")

    return _apply_mask(result, water_mask, apply_mask)


def _gee_viz_params(parameter: str) -> Dict[str, Any]:
    gad_palette = ["#08306b", "#2171b5", "#6baed6", "#c7e9b4", "#fee08b", "#fdae61", "#f46d43", "#d73027"]
    default_palette = ["#0000FF", "#00FFFF", "#00FF00", "#FFFF00", "#FF0000"]
    if parameter.startswith("True Color"):
        return {"bands": ["B4", "B3", "B2"], "min": 0, "max": 3000}
    if parameter.startswith("Water Surface Temperature"):
        return {"min": 5, "max": 35, "palette": ["#000080", "#0000FF", "#00FFFF", "#FFFF00", "#FF0000", "#800000"]}
    if parameter.startswith("NDWI"):
        return {"min": -0.5, "max": 0.8, "palette": ["#8B4513", "#FFFF00", "#00FFFF", "#0000FF"]}
    if parameter.startswith("NDTI"):
        return {"min": -0.2, "max": 0.3, "palette": ["#0000FF", "#00FFFF", "#FFFF00", "#FF8000", "#FF0000"]}
    if parameter.startswith("TSM"):
        return {"min": 0, "max": 100, "palette": ["#0000FF", "#00FFFF", "#FFFF00", "#FF8000", "#FF0000"]}
    if parameter.startswith("Chl-a (Gadouras Autumn)"):
        return {"min": 10, "max": 35, "palette": gad_palette}
    if parameter.startswith("Chl-a (Gadouras"):
        return {"min": 0, "max": 35, "palette": gad_palette}
    if parameter.startswith("NDCI (Gadouras scale)"):
        return {"min": -0.35, "max": 0.15, "palette": ["#2166ac", "#92c5de", "#f7f7f7", "#fddbc7", "#b2182b"]}
    if parameter.startswith("Anomaly Detection"):
        return {"min": 0, "max": 1, "palette": ["#00FF00", "#FF0000"]}
    if parameter.startswith("Water Turbidity Classes"):
        return {"min": 1, "max": 5, "palette": ["#0000FF", "#00FFFF", "#FFFF00", "#FF8000", "#FF0000"]}
    if parameter.startswith("Algae Bloom Detection"):
        return {"min": 0, "max": 3, "palette": default_palette}
    if parameter.startswith("Chl-a"):
        return {"min": 0, "max": 50, "palette": default_palette}
    if parameter.startswith("NDCI"):
        return {"min": -0.2, "max": 0.6, "palette": default_palette}
    return {"min": 0, "max": 1, "palette": default_palette}


def _get_gee_collection(parameter: str, start_date, end_date, bounds, cloud_pct: int = 35):
    if "Temperature" in parameter:
        l8 = ee.ImageCollection("LANDSAT/LC08/C02/T1_L2").filterBounds(bounds).filterDate(start_date, end_date).map(_mask_l8_sr)
        l9 = ee.ImageCollection("LANDSAT/LC09/C02/T1_L2").filterBounds(bounds).filterDate(start_date, end_date).map(_mask_l8_sr)
        return ee.ImageCollection(l8.merge(l9)).sort("system:time_start")
    return (
        ee.ImageCollection("COPERNICUS/S2_SR_HARMONIZED")
        .filterBounds(bounds)
        .filterDate(start_date, end_date)
        .filter(ee.Filter.lt("CLOUDY_PIXEL_PERCENTAGE", cloud_pct))
        .sort("system:time_start")
    )


def _normalize_feature_geojson(draw_obj: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    if not draw_obj:
        return None
    if draw_obj.get("type") == "Feature" and "geometry" in draw_obj:
        return draw_obj
    if "geometry" in draw_obj and "type" in draw_obj["geometry"]:
        return {"type": "Feature", "geometry": draw_obj["geometry"], "properties": draw_obj.get("properties", {})}
    if "type" in draw_obj and "coordinates" in draw_obj:
        return {"type": "Feature", "geometry": draw_obj, "properties": {}}
    return None


def _geojson_to_ee_geometry(feature_geojson: Dict[str, Any]):
    geom = feature_geojson.get("geometry", feature_geojson)
    gtype = geom.get("type")
    coords = geom.get("coordinates")
    if gtype == "Polygon":
        return ee.Geometry.Polygon(coords), gtype
    if gtype == "MultiPolygon":
        return ee.Geometry.MultiPolygon(coords), gtype
    if gtype == "LineString":
        return ee.Geometry.LineString(coords), gtype
    if gtype == "MultiLineString":
        return ee.Geometry.MultiLineString(coords), gtype
    if gtype == "Point":
        return ee.Geometry.Point(coords), gtype
    raise ValueError(f"Unsupported geometry type: {gtype}")


def _ee_tile_url(image, viz_params: Dict[str, Any]) -> str:
    map_id = image.getMapId(viz_params)
    return map_id["tile_fetcher"].url_format


def _build_gee_map(mode: str, draw_feature: Optional[Dict[str, Any]], layer_url: Optional[str], layer_name: str):
    if folium is None:
        return None
    m = folium.Map(location=[36.172, 27.988], zoom_start=11, control_scale=True, tiles=None)
    # Keep satellite as default basemap for this tab.
    folium.TileLayer("OpenStreetMap", name="OpenStreetMap", overlay=False, control=True, show=False).add_to(m)
    folium.TileLayer(
        tiles="https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}",
        attr="Esri",
        name="Satellite",
        overlay=False,
        control=True,
        show=True,
    ).add_to(m)

    if layer_url:
        folium.raster_layers.TileLayer(
            tiles=layer_url,
            attr="Google Earth Engine",
            name=layer_name,
            overlay=True,
            control=True,
            show=True,
            opacity=0.9,
        ).add_to(m)

    if draw_feature:
        folium.GeoJson(draw_feature, name="Selected geometry").add_to(m)

    if Draw is not None:
        draw_options = {
            "polyline": mode == "Transect",
            "polygon": mode != "Transect",
            "rectangle": mode != "Transect",
            "circle": False,
            "marker": False,
            "circlemarker": False,
        }
        Draw(
            export=False,
            draw_options=draw_options,
            edit_options={"edit": True, "remove": True},
        ).add_to(m)
    folium.LayerControl(position="topright").add_to(m)
    return m


def _line_points(line_geom, num_points: int):
    n = max(int(num_points), 2)

    # Earth Engine Python Geometry currently does not expose .interpolate() reliably.
    # Build transect points client-side from line coordinates and return as EE features.
    coords = line_geom.coordinates().getInfo() or []
    if not isinstance(coords, list) or len(coords) < 2:
        return ee.FeatureCollection([])

    def _haversine_m(p1, p2):
        lon1, lat1 = float(p1[0]), float(p1[1])
        lon2, lat2 = float(p2[0]), float(p2[1])
        r = 6371000.0
        phi1 = math.radians(lat1)
        phi2 = math.radians(lat2)
        dphi = math.radians(lat2 - lat1)
        dlambda = math.radians(lon2 - lon1)
        a = math.sin(dphi / 2.0) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2.0) ** 2
        return 2.0 * r * math.atan2(math.sqrt(a), math.sqrt(max(0.0, 1.0 - a)))

    seg_lengths = []
    total_length = 0.0
    for i in range(len(coords) - 1):
        seg_len = _haversine_m(coords[i], coords[i + 1])
        seg_lengths.append(seg_len)
        total_length += seg_len

    if total_length <= 0:
        pt = ee.Feature(ee.Geometry.Point(coords[0]), {"distance_m": 0.0, "index": 0})
        return ee.FeatureCollection([pt])

    targets = [i * total_length / (n - 1) for i in range(n)]
    features = []
    cum_len = 0.0
    seg_idx = 0

    for idx, tdist in enumerate(targets):
        while seg_idx < len(seg_lengths) - 1 and tdist > cum_len + seg_lengths[seg_idx]:
            cum_len += seg_lengths[seg_idx]
            seg_idx += 1

        seg_len = seg_lengths[seg_idx]
        if seg_len <= 0:
            frac = 0.0
        else:
            frac = (tdist - cum_len) / seg_len
            frac = max(0.0, min(1.0, frac))

        p1 = coords[seg_idx]
        p2 = coords[seg_idx + 1]
        lon = float(p1[0]) + (float(p2[0]) - float(p1[0])) * frac
        lat = float(p1[1]) + (float(p2[1]) - float(p1[1])) * frac

        features.append(
            ee.Feature(
                ee.Geometry.Point([lon, lat]),
                {"distance_m": float(tdist), "index": int(idx)},
            )
        )

    return ee.FeatureCollection(features)


def _run_transect_analysis(collection, line_geom, parameter: str, num_points: int, max_images: Optional[int] = None):
    image_count = int(collection.size().getInfo() or 0)
    if image_count == 0:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), None, 0, 0

    latest_img = ee.Image(collection.sort("system:time_start", False).first())
    is_temperature = "Temperature" in str(parameter)
    latest_date = ee.Date(latest_img.get("system:time_start")).format("YYYY-MM-dd").getInfo()
    latest_processed = _process_gee_image(latest_img, parameter, apply_mask=True)
    band_name = ee.String(latest_processed.bandNames().get(0)).getInfo()

    def _profile_from_sampled(sampled_fc, value_key: str) -> pd.DataFrame:
        sampled_info = sampled_fc.getInfo() or {}
        profile_rows = []
        for feat in sampled_info.get("features", []):
            props = feat.get("properties", {}) or {}
            val = props.get(value_key, None)
            if val is None:
                val = props.get("first", None)
            if val is None:
                for k, v in props.items():
                    if k in ("distance_m", "index", "system:index"):
                        continue
                    if isinstance(v, (int, float)):
                        val = v
                        break
            profile_rows.append(
                {
                    "distance_m": props.get("distance_m", np.nan),
                    "value": val if val is not None else np.nan,
                    "index": props.get("index", np.nan),
                }
            )
        profile_df_local = pd.DataFrame(profile_rows)
        if profile_df_local.empty:
            return pd.DataFrame(columns=["distance_m", "value", "index"])
        profile_df_local["distance_m"] = pd.to_numeric(profile_df_local["distance_m"], errors="coerce")
        profile_df_local["value"] = pd.to_numeric(profile_df_local["value"], errors="coerce")
        profile_df_local["index"] = pd.to_numeric(profile_df_local["index"], errors="coerce")
        profile_df_local = profile_df_local.dropna(subset=["distance_m", "value"]).sort_values(["distance_m", "index"])
        return profile_df_local

    points_fc = _line_points(line_geom, num_points)
    sampled = latest_processed.reduceRegions(collection=points_fc, reducer=ee.Reducer.first(), scale=20)
    profile_df = _profile_from_sampled(sampled, band_name)

    # Fallback: if masking removed all transect values, retry without water mask.
    if profile_df.empty and is_temperature:
        latest_processed_nomask = _process_gee_image(latest_img, parameter, apply_mask=False)
        band_name_nomask = ee.String(latest_processed_nomask.bandNames().get(0)).getInfo()
        sampled_nomask = latest_processed_nomask.reduceRegions(collection=points_fc, reducer=ee.Reducer.first(), scale=20)
        profile_df = _profile_from_sampled(sampled_nomask, band_name_nomask)

    if max_images is None:
        use_count = image_count
    else:
        try:
            use_count = min(int(max_images), image_count)
        except Exception:
            use_count = image_count
    use_count = max(1, use_count)
    img_list = collection.sort("system:time_start").toList(use_count)
    ts_rows = []
    hov_parts = []
    for i in range(use_count):
        img = ee.Image(img_list.get(i))
        date_str = ee.Date(img.get("system:time_start")).format("YYYY-MM-dd").getInfo()
        proc = _process_gee_image(img, parameter, apply_mask=True)
        bn = ee.String(proc.bandNames().get(0)).getInfo()
        mean_val = proc.reduceRegion(
            reducer=ee.Reducer.mean(),
            geometry=line_geom.buffer(60),
            scale=20,
            maxPixels=1_000_000_000,
        ).get(bn).getInfo()
        if mean_val is None and is_temperature:
            proc_nomask = _process_gee_image(img, parameter, apply_mask=False)
            bn_nomask = ee.String(proc_nomask.bandNames().get(0)).getInfo()
            mean_val = proc_nomask.reduceRegion(
                reducer=ee.Reducer.mean(),
                geometry=line_geom.buffer(60),
                scale=20,
                maxPixels=1_000_000_000,
            ).get(bn_nomask).getInfo()
        ts_rows.append({"date": pd.to_datetime(date_str), "mean": mean_val})

        # Build Hovmoller rows (distance x time). Fallback to no-mask if masked sampling is empty.
        sampled_img = proc.reduceRegions(collection=points_fc, reducer=ee.Reducer.first(), scale=20)
        prof_img = _profile_from_sampled(sampled_img, bn)
        if prof_img.empty and is_temperature:
            proc_nomask = _process_gee_image(img, parameter, apply_mask=False)
            bn_nomask = ee.String(proc_nomask.bandNames().get(0)).getInfo()
            sampled_img_nomask = proc_nomask.reduceRegions(collection=points_fc, reducer=ee.Reducer.first(), scale=20)
            prof_img = _profile_from_sampled(sampled_img_nomask, bn_nomask)
        if not prof_img.empty:
            prof_img = prof_img[["distance_m", "value"]].copy()
            prof_img["date"] = pd.to_datetime(date_str)
            hov_parts.append(prof_img)

    ts_df = pd.DataFrame(ts_rows).dropna(subset=["mean"]).sort_values("date")
    hov_df = pd.concat(hov_parts, ignore_index=True) if hov_parts else pd.DataFrame(columns=["date", "distance_m", "value"])
    return profile_df, ts_df, hov_df, latest_date, use_count, image_count


def _lake_water_mask(image, method: str, threshold: float):
    ndwi = image.normalizedDifference(["B3", "B8"]).rename("ndwi")
    mndwi = image.normalizedDifference(["B3", "B11"]).rename("mndwi")
    awei = image.expression(
        "4*(GREEN-SWIR1)-(0.25*NIR+2.75*SWIR2)",
        {
            "GREEN": image.select("B3"),
            "NIR": image.select("B8"),
            "SWIR1": image.select("B11"),
            "SWIR2": image.select("B12"),
        },
    ).rename("awei")
    ndvi = image.normalizedDifference(["B8", "B4"]).rename("ndvi")

    if method == "NDWI Standard":
        wm = ndwi.gt(threshold)
    elif method == "MNDWI Enhanced":
        wm = mndwi.gt(threshold)
    elif method == "AWEInsh":
        wm = awei.gt(threshold * 1000)
    elif method == "Multi-Index Fusion":
        wm = ndwi.gt(threshold).Or(mndwi.gt(threshold)).And(ndvi.lt(0.3)).And(awei.gt(0))
    elif method == "Simple Threshold":
        wm = image.select("B8").lt(1000)
    else:
        wm = mndwi.gt(threshold - 0.1).And(ndvi.lt(0.3)).Or(ndwi.gt(threshold).And(awei.gt(0)))
    return wm.focal_min(1).focal_max(1).selfMask().rename("water")


def _run_lake_height_analysis(collection, polygon_geom, method: str, threshold: float, max_images: int = 24):
    dem = ee.ImageCollection("COPERNICUS/DEM/GLO30").select("DEM").mosaic()
    image_count = int(collection.size().getInfo() or 0)
    if image_count == 0:
        return pd.DataFrame()
    use_count = min(max_images, image_count)
    img_list = collection.sort("system:time_start").toList(use_count)
    rows = []
    for i in range(use_count):
        img = ee.Image(img_list.get(i)).select(["B2", "B3", "B4", "B8", "B11", "B12"])
        date_str = ee.Date(img.get("system:time_start")).format("YYYY-MM-dd").getInfo()
        water = _lake_water_mask(img, method, threshold)
        area = water.multiply(ee.Image.pixelArea()).reduceRegion(
            reducer=ee.Reducer.sum(),
            geometry=polygon_geom,
            scale=20,
            maxPixels=1_000_000_000,
        ).get("water")
        wb = water.unmask(0)
        edge = wb.focal_max(1).subtract(wb).gt(0)
        elev = dem.updateMask(edge).reduceRegion(
            reducer=ee.Reducer.median(),
            geometry=polygon_geom,
            scale=30,
            maxPixels=1_000_000_000,
        ).get("DEM")
        area_val = ee.Number(area).divide(1e6).getInfo() if area is not None else None
        elev_val = ee.Number(elev).getInfo() if elev is not None else None
        rows.append(
            {
                "date": pd.to_datetime(date_str),
                "water_area_km2": area_val,
                "shoreline_elev_m": elev_val,
            }
        )
    return pd.DataFrame(rows).dropna(how="all").sort_values("date")


tab_level, tab_map, tab_gee, tab_ts, tab_3d, tab_depth, tab_compare, tab_raw = st.tabs([
    "Στάθμη",
    "Map",
    "GEE Explorer v4",
    "Time Series",
    "3D Maps",
    "Depth Profiles",
    "Compare Points",
    "Data",
])
# ══════════════════════════════════════════════════════════════
# TAB 1: LEVEL
# ══════════════════════════════════════════════════════════════
with tab_level:
    render_level_tab()

# ══════════════════════════════════════════════════════════════
# TAB 2: MAP
# ══════════════════════════════════════════════════════════════
with tab_map:
    st.subheader("Θέσεις Σημείων Δειγματοληψίας")
    
    fig_map = build_sampling_map(
        height=550,
        title="Ταμιευτήρας Φράγματος Γαδουρά - Θέσεις Δειγματοληψίας",
        color_param=map_param_key,
        sel_date=map_date_value,
        sel_depth=map_depth_value,
    )
    st.plotly_chart(fig_map, use_container_width=True, theme=None)
    
    map_table_df = _build_map_dataset(
        color_param=map_param_key,
        sel_date=map_date_value,
        sel_depth=map_depth_value,
    ).copy()
    map_table_df["point_number"] = map_table_df["point_id"]

    if map_param_key == "__counts__":
        table_view = map_table_df.rename(columns={
            "point_number": "#",
            "point_label": "Σημείο",
            "lat": "Γεωγρ. Πλάτος",
            "lon": "Γεωγρ. Μήκος",
            "measurement_count": "Μετρήσεις",
        })[["#", "Σημείο", "Γεωγρ. Πλάτος", "Γεωγρ. Μήκος", "Μετρήσεις"]]
    else:
        map_table_df[map_param_choice] = map_table_df["param_value"].round(3)
        table_view = map_table_df.rename(columns={
            "point_number": "#",
            "point_label": "Σημείο",
            "lat": "Γεωγρ. Πλάτος",
            "lon": "Γεωγρ. Μήκος",
            "measurement_count": "Μετρήσεις",
        })[["#", "Σημείο", "Γεωγρ. Πλάτος", "Γεωγρ. Μήκος", "Μετρήσεις", map_param_choice]]

    st.dataframe(table_view, use_container_width=True, hide_index=True)

# TAB 2: GEE EXPLORER
with tab_gee:
    st.subheader("Interactive Water Quality Explorer (GEE v4)")
    st.caption(
        "Modes: ROI, Transect, Lake Height. Includes Gadouras calibrated chlorophyll models. "
        "Draw geometry on the map, choose settings, then run analysis."
    )

    init_col_l, init_col_r = st.columns([2.1, 1.2])
    with init_col_l:
        gee_project_input = st.text_input(
            "EE Project (optional)",
            value=os.getenv("EE_PROJECT", DEFAULT_EE_PROJECT),
            key="gee_project_input_v4",
            help="If set, we call ee.Initialize(project=...). Leave empty to use your EE default project.",
        )
    with init_col_r:
        gee_use_auto_sa = st.checkbox(
            "Use local service account key",
            value=False,
            key="gee_use_auto_sa_v4",
            help="Auto-detects local ee-*.json key files in this app folder.",
        )

    svc_email_input = ""
    svc_key_payload = None
    with st.expander("Service account (optional)", expanded=False):
        svc_email_input = st.text_input(
            "Service account email",
            value=os.getenv("EE_SERVICE_ACCOUNT", ""),
            key="gee_sa_email_v4",
        )
        svc_key_file = st.file_uploader(
            "Service account key JSON",
            type="json",
            key="gee_sa_key_upload_v4",
            help="Uploaded in-memory only for this Streamlit session.",
        )
        if svc_key_file is not None:
            try:
                svc_key_payload = svc_key_file.getvalue().decode("utf-8")
                json.loads(svc_key_payload)
            except Exception as exc:
                svc_key_payload = None
                st.warning(f"Invalid service-account JSON: {exc}")

    ee_ready, ee_msg = _init_ee(
        project_id=(gee_project_input or "").strip() or None,
        svc_account_email=(svc_email_input or "").strip() or None,
        svc_key_json=svc_key_payload,
        allow_auto_service_account=gee_use_auto_sa,
    )

    if not ee_ready:
        st.warning("Google Earth Engine is not initialized in this environment.")
        st.code("earthengine authenticate && earthengine set_project <your-project-id>", language="bash")
        if ee_msg:
            st.caption(f"EE init details: {ee_msg}")
    elif folium is None or st_folium is None:
        st.error("Missing map dependencies (`folium` and `streamlit-folium`).")
    else:
        if "gee_draw_feature" not in st.session_state:
            st.session_state["gee_draw_feature"] = None
        if "gee_layer_url" not in st.session_state:
            st.session_state["gee_layer_url"] = None
        if "gee_layer_name" not in st.session_state:
            st.session_state["gee_layer_name"] = "Processed layer"
        if "gee_notice" not in st.session_state:
            st.session_state["gee_notice"] = None
        if st.session_state.get("gee_notice"):
            st.success(st.session_state["gee_notice"])
            st.session_state["gee_notice"] = None

        c_mode, c_dates = st.columns([1.2, 2.2])
        with c_mode:
            gee_mode = st.selectbox("Mode", ["ROI", "Transect", "Lake Height"], key="gee_mode_select")
        with c_dates:
            d1, d2 = st.columns(2)
            with d1:
                gee_start = st.date_input("Start date", value=pd.to_datetime("2023-01-01"), key="gee_start")
            with d2:
                gee_end = st.date_input("End date", value=pd.to_datetime("today"), key="gee_end")

        map_col, ctrl_col = st.columns([1.8, 1.2], gap='large')
        with map_col:
            gee_map = _build_gee_map(
                mode=gee_mode,
                draw_feature=st.session_state.get("gee_draw_feature"),
                layer_url=st.session_state.get("gee_layer_url"),
                layer_name=st.session_state.get("gee_layer_name", "Processed layer"),
            )
            map_data = st_folium(
                gee_map,
                key="gee_map_canvas",
                height=620,
                width=None,
                returned_objects=["last_active_drawing", "all_drawings"],
            )
            if map_data:
                last_draw = map_data.get("last_active_drawing")
                all_draws = map_data.get("all_drawings") or []
                normalized = _normalize_feature_geojson(last_draw) if last_draw else None
                if normalized is None and len(all_draws) > 0:
                    normalized = _normalize_feature_geojson(all_draws[-1])
                if normalized is not None:
                    st.session_state["gee_draw_feature"] = normalized

        with ctrl_col:
            gee_param = st.selectbox("Parameter", options=GEE_PARAMETER_OPTIONS, key="gee_param")
            gee_cloud = st.slider("Max cloud (%)", 0, 100, 35, 5, key="gee_cloud")
            gee_max_images = st.slider("Max images to process", 3, 40, 12, 1, key="gee_max_images")
            gee_use_all_transect = st.checkbox("Transect: use all available images", value=True, key="gee_transect_use_all")
            gee_points = st.slider("Transect points", 10, 200, 50, 5, key="gee_points")
            gee_lake_method = st.selectbox("Lake water method", options=GEE_LAKE_METHODS, key="gee_lake_method")
            gee_lake_thr = st.slider("Lake water sensitivity", -0.5, 0.5, 0.0, 0.05, key="gee_lake_thr")

            b1, b2 = st.columns(2)
            with b1:
                if st.button("Clear drawing", key="gee_clear_drawing", use_container_width=True):
                    st.session_state["gee_draw_feature"] = None
                    st.session_state["gee_layer_url"] = None
                    st.session_state["gee_layer_name"] = "Processed layer"
            with b2:
                run_gee = st.button("Run GEE analysis", key="gee_run_btn", type="primary", use_container_width=True)

        refresh_map_after_run = False
        if run_gee:
            draw_feature = st.session_state.get("gee_draw_feature")
            if draw_feature is None:
                st.error("Draw a geometry first on the map.")
            elif gee_start > gee_end:
                st.error("Start date must be before end date.")
            else:
                try:
                    geometry, geom_type = _geojson_to_ee_geometry(draw_feature)
                except Exception as exc:
                    st.error(f"Invalid geometry: {exc}")
                    geometry = None
                    geom_type = None

                if geometry is not None:
                    with st.spinner("Running Earth Engine analysis..."):
                        try:
                            if gee_mode == "ROI":
                                if geom_type not in ("Polygon", "MultiPolygon"):
                                    st.error("ROI mode requires Polygon/Rectangle drawing.")
                                else:
                                    col = _get_gee_collection(gee_param, str(gee_start), str(gee_end), geometry, cloud_pct=gee_cloud)
                                    count = int(col.size().getInfo() or 0)
                                    if count == 0:
                                        st.warning("No images found for the selected period/geometry.")
                                    else:
                                        latest = ee.Image(col.sort("system:time_start", False).first())
                                        latest_date = ee.Date(latest.get("system:time_start")).format("YYYY-MM-dd").getInfo()
                                        proc = _process_gee_image(latest, gee_param, apply_mask=True).clip(geometry)
                                        viz = _gee_viz_params(gee_param)
                                        st.session_state["gee_layer_url"] = _ee_tile_url(proc, viz)
                                        st.session_state["gee_layer_name"] = f"{gee_param} ({latest_date})"
                                        st.session_state["gee_notice"] = f"Loaded {count} images. Showing latest: {latest_date}"
                                        refresh_map_after_run = True

                                        if "bands" not in viz:
                                            band = ee.String(proc.bandNames().get(0)).getInfo()
                                            mn_mx = proc.reduceRegion(
                                                reducer=ee.Reducer.minMax(),
                                                geometry=geometry,
                                                scale=20,
                                                maxPixels=1_000_000_000,
                                            ).getInfo() or {}
                                            mean_v = proc.reduceRegion(
                                                reducer=ee.Reducer.mean(),
                                                geometry=geometry,
                                                scale=20,
                                                maxPixels=1_000_000_000,
                                            ).getInfo() or {}
                                            c1, c2, c3 = st.columns(3)
                                            c1.metric("ROI Mean", f"{float(mean_v.get(band, np.nan)):.3f}" if band in mean_v else "n/a")
                                            c2.metric("ROI Min", f"{float(mn_mx.get(f'{band}_min', np.nan)):.3f}" if f"{band}_min" in mn_mx else "n/a")
                                            c3.metric("ROI Max", f"{float(mn_mx.get(f'{band}_max', np.nan)):.3f}" if f"{band}_max" in mn_mx else "n/a")

                            elif gee_mode == "Transect":
                                if geom_type not in ("LineString", "MultiLineString"):
                                    st.error("Transect mode requires a line geometry.")
                                else:
                                    if geom_type == "MultiLineString":
                                        first_line = ee.List(geometry.coordinates()).get(0)
                                        geometry = ee.Geometry.LineString(first_line)
                                    col = _get_gee_collection(gee_param, str(gee_start), str(gee_end), geometry, cloud_pct=gee_cloud)
                                    transect_limit = None if gee_use_all_transect else gee_max_images
                                    profile_df, ts_df, hov_df, latest_date, used_images, total_images = _run_transect_analysis(
                                        collection=col,
                                        line_geom=geometry,
                                        parameter=gee_param,
                                        num_points=gee_points,
                                        max_images=transect_limit,
                                    )
                                    if profile_df.empty:
                                        st.warning("No transect values were computed for the selected period.")
                                    else:
                                        if used_images < total_images:
                                            st.info(
                                                f"Transect processed {used_images}/{total_images} images "
                                                f"(limited by Max images to process = {gee_max_images})."
                                            )
                                        else:
                                            st.caption(f"Transect processed all {total_images} available images.")
                                        st.session_state["gee_layer_name"] = f"{gee_param} profile ({latest_date})"
                                        latest_img = ee.Image(col.sort("system:time_start", False).first())
                                        proc = _process_gee_image(latest_img, gee_param, apply_mask=True).clip(geometry.buffer(200))
                                        viz = _gee_viz_params(gee_param)
                                        st.session_state["gee_layer_url"] = _ee_tile_url(proc, viz)

                                        st.markdown("**Transect profile (latest image)**")
                                        fig_prof = px.line(
                                            profile_df,
                                            x="distance_m",
                                            y="value",
                                            markers=True,
                                            labels={"distance_m": "Distance along line (m)", "value": gee_param},
                                        )
                                        fig_prof.update_layout(height=320)
                                        st.plotly_chart(fig_prof, use_container_width=True, theme=None)

                                        if not hov_df.empty:
                                            st.markdown("**Transect Hovmoller (distance × time)**")
                                            hov_plot = hov_df.copy()
                                            hov_plot["distance_m"] = pd.to_numeric(hov_plot["distance_m"], errors="coerce").round(2)
                                            hov_plot["value"] = pd.to_numeric(hov_plot["value"], errors="coerce")
                                            hov_plot = hov_plot.dropna(subset=["date", "distance_m", "value"])
                                            if not hov_plot.empty:
                                                hov_viz = _gee_viz_params(gee_param)
                                                hov_min = hov_viz.get("min", float(hov_plot["value"].min()))
                                                hov_max = hov_viz.get("max", float(hov_plot["value"].max()))
                                                piv = (
                                                    hov_plot.pivot_table(
                                                        index="distance_m",
                                                        columns="date",
                                                        values="value",
                                                        aggfunc="mean",
                                                    )
                                                    .sort_index()
                                                    .sort_index(axis=1)
                                                )
                                                if piv.shape[0] >= 1 and piv.shape[1] >= 2:
                                                    fig_hov = px.imshow(
                                                        piv,
                                                        aspect="auto",
                                                        color_continuous_scale="Turbo",
                                                        zmin=hov_min,
                                                        zmax=hov_max,
                                                        labels={
                                                            "x": "Date",
                                                            "y": "Distance along line (m)",
                                                            "color": gee_param,
                                                        },
                                                    )
                                                    fig_hov.update_layout(height=360)
                                                    st.plotly_chart(fig_hov, use_container_width=True, theme=None)

                                        if not ts_df.empty:
                                            st.caption(f"Transect mean valid values: {len(ts_df)}/{used_images} processed images.")
                                            st.markdown("**Transect mean over time**")
                                            fig_ts_gee = px.line(
                                                ts_df,
                                                x="date",
                                                y="mean",
                                                markers=True,
                                                labels={"date": "Date", "mean": f"Mean {gee_param}"},
                                            )
                                            fig_ts_gee.update_layout(height=320)
                                            st.plotly_chart(fig_ts_gee, use_container_width=True, theme=None)

                            else:
                                if geom_type not in ("Polygon", "MultiPolygon"):
                                    st.error("Lake Height mode requires Polygon/Rectangle drawing.")
                                else:
                                    lake_col = (
                                        ee.ImageCollection("COPERNICUS/S2_SR_HARMONIZED")
                                        .filterBounds(geometry)
                                        .filterDate(str(gee_start), str(gee_end))
                                        .filter(ee.Filter.lt("CLOUDY_PIXEL_PERCENTAGE", gee_cloud))
                                        .sort("system:time_start")
                                    )
                                    lake_df = _run_lake_height_analysis(
                                        collection=lake_col,
                                        polygon_geom=geometry,
                                        method=gee_lake_method,
                                        threshold=float(gee_lake_thr),
                                        max_images=gee_max_images,
                                    )
                                    if lake_df.empty:
                                        st.warning("No valid lake-height observations were produced.")
                                    else:
                                        st.success(f"Computed {len(lake_df)} lake observations.")
                                        fig_area = px.line(
                                            lake_df,
                                            x="date",
                                            y="water_area_km2",
                                            markers=True,
                                            labels={"date": "Date", "water_area_km2": "Water area (km²)"},
                                        )
                                        fig_area.update_layout(height=310)
                                        st.plotly_chart(fig_area, use_container_width=True, theme=None)

                                        fig_elev = px.line(
                                            lake_df,
                                            x="date",
                                            y="shoreline_elev_m",
                                            markers=True,
                                            labels={"date": "Date", "shoreline_elev_m": "Shoreline elevation (m)"},
                                        )
                                        fig_elev.update_layout(height=310)
                                        st.plotly_chart(fig_elev, use_container_width=True, theme=None)

                                        latest_lake = ee.Image(lake_col.sort("system:time_start", False).first()).select(
                                            ["B2", "B3", "B4", "B8", "B11", "B12"]
                                        )
                                        wm = _lake_water_mask(latest_lake, gee_lake_method, float(gee_lake_thr)).clip(geometry)
                                        st.session_state["gee_layer_url"] = _ee_tile_url(
                                            wm,
                                            {"min": 0, "max": 1, "palette": ["#1d4ed8"]},
                                        )
                                        st.session_state["gee_layer_name"] = "Lake water mask"

                        except Exception as exc:
                            st.error(f"GEE analysis failed: {exc}")
        if refresh_map_after_run:
            st.rerun()

# TAB 2: TIME SERIES
# ══════════════════════════════════════════════════════════════
with tab_ts:
    st.subheader("Χρονοσειρές Παραμέτρων")
    
    col_a, col_b, col_c = st.columns(1) if _is_mobile else st.columns([2, 2, 2])
    
    with col_a:
        param = st.selectbox(
            "Παράμετρος",
            options=[p for p in COL_MAP if df[p].notna().any()],
            key="ts_param"
        )
    
    all_points = sorted(df["point"].unique())
    with col_b:
        sel_points = st.multiselect(
            "Σημεία Δειγματοληψίας",
            options=all_points,
            default=all_points,
            format_func=lambda x: f"Σημείο {x}",
            key="ts_points"
        )
    
    all_depths = [d for d in DEPTH_ORDER if d in df["depth"].unique()]
    with col_c:
        sel_depth = st.selectbox(
            "Βάθος",
            options=["Όλα"] + all_depths,
            key="ts_depth"
        )
    
    if not sel_points:
        st.warning("Επιλέξτε τουλάχιστον ένα σημείο.")
    else:
        filt = df[df["point"].isin(sel_points)].copy()
        if sel_depth != "Όλα":
            filt = filt[filt["depth"] == sel_depth]
        
        # Aggregate (mean per date/point when multiple depths)
        grp = filt.groupby(["date", "point"])[param].mean().reset_index()
        grp_nonnull = grp.dropna(subset=[param])
        
        if grp_nonnull.empty:
            st.info("Δεν υπάρχουν διαθέσιμα δεδομένα για τον συνδυασμό που επιλέξατε.")
        else:
            fig = go.Figure()
            colors = _PLT_SERIES
            for idx, pt in enumerate(sel_points):
                sub = grp_nonnull[grp_nonnull["point"] == pt].sort_values("date")
                if sub.empty:
                    continue
                clr = colors[idx % len(colors)]
                fig.add_trace(go.Scatter(
                    x=sub["date"], y=sub[param],
                    mode="lines+markers",
                    name=f"Σημείο {pt}",
                    line=dict(color=clr, width=2),
                    marker=dict(size=8, color=clr, line=dict(width=1.5, color=_PLT_MARKER_BORDER)),
                    hovertemplate=f"<b>Σημείο {pt}</b><br>Ημ/νία: %{{x|%d/%m/%Y}}<br>{param}: %{{y:.3f}}<extra></extra>"
                ))
            
            fig.update_layout(
                title=dict(text=f"<b>{param}</b>  —  Χρονοσειρά", font=dict(size=15)),
                xaxis=dict(title="Ημερομηνία", tickformat="%d/%m/%Y", tickangle=-30),
                yaxis=dict(title=param),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                hovermode="x unified",
                height=480,
                plot_bgcolor=_PLT_BG,
                paper_bgcolor=_PLT_PAPER,
            )
            fig.update_layout(
                font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                hoverlabel=dict(bgcolor=_PLT_HOVER_BG, font_color=_PLT_HOVER_FONT, bordercolor=_PLT_LEGEND_BORDER),
                legend=dict(
                    orientation="h",
                    yanchor="bottom",
                    y=1.02,
                    xanchor="right",
                    x=1,
                    bgcolor=_PLT_LEGEND_BG,
                    bordercolor=_PLT_LEGEND_BORDER,
                    borderwidth=1,
                    font=dict(color=_PLT_TICK)
                ),
                plot_bgcolor=_PLT_BG,
                paper_bgcolor=_PLT_PAPER,
            )
            fig.update_xaxes(
                showgrid=True,
                gridcolor=_PLT_GRID,
                tickfont=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                title_font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                linecolor=_PLT_LINE
            )
            fig.update_yaxes(
                showgrid=True,
                gridcolor=_PLT_GRID,
                tickfont=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                title_font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                linecolor=_PLT_LINE
            )
            st.plotly_chart(fig, use_container_width=True, theme=None)
        
        # Heatmap: param across dates & points
        if not grp.empty:
            st.markdown("#### 🌡️ Θερμικός Χάρτης")
            pivot = grp.pivot(index="point", columns="date", values=param)
            heat_dates = sorted(pd.to_datetime(filt["date"]).dropna().unique())
            if heat_dates:
                pivot = pivot.reindex(index=sel_points, columns=heat_dates)
            pivot.columns = [pd.to_datetime(d).strftime("%d/%m/%Y") for d in pivot.columns]
            pivot.index = [f"Σ{p}" for p in pivot.index]
            
            fig_h = px.imshow(
                pivot,
                labels=dict(x="Ημερομηνία", y="Σημείο", color=param),
                color_continuous_scale="RdYlBu_r",
                aspect="auto",
                title=f"Heatmap: {param}",
                height=350
            )
            fig_h.update_layout(
                plot_bgcolor=_PLT_BG,
                paper_bgcolor=_PLT_PAPER,
                font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                margin=dict(t=55, b=95, l=70, r=20),
                coloraxis_colorbar=dict(
                    outlinecolor=_PLT_LINE,
                    outlinewidth=1,
                    tickfont=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                    title_font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                ),
            )
            fig_h.update_xaxes(
                tickangle=-25,
                automargin=True,
                showgrid=False,
                linecolor=_PLT_LINE,
                tickfont=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                title_font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                title_standoff=18,
            )
            fig_h.update_yaxes(
                automargin=True,
                showgrid=False,
                linecolor=_PLT_LINE,
                tickfont=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                title_font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                title_standoff=10,
            )
            st.plotly_chart(fig_h, use_container_width=True, theme=None)

            # Timeseries of per-date average across all selected points
            avg_ts = grp_nonnull.groupby("date")[param].mean().reset_index().sort_values("date")
            if not avg_ts.empty:
                st.markdown("#### Average Per Date (All Selected Points)")
                fig_avg = go.Figure()
                fig_avg.add_trace(go.Scatter(
                    x=avg_ts["date"],
                    y=avg_ts[param],
                    mode="lines+markers",
                    name="Average",
                    line=dict(color=_PLT_SERIES[1], width=3),
                    marker=dict(size=8, color=_PLT_SERIES[1], line=dict(width=1, color=_PLT_MARKER_BORDER)),
                    hovertemplate="Date: %{x|%d/%m/%Y}<br>Average: %{y:.3f}<extra></extra>",
                ))
                fig_avg.update_layout(
                    title=f"Average {param} per Date",
                    xaxis=dict(title="Ημερομηνία", tickformat="%d/%m/%Y", tickangle=-30),
                    yaxis=dict(title=param),
                    height=340,
                    plot_bgcolor=_PLT_BG,
                    paper_bgcolor=_PLT_PAPER,
                    font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                    margin=dict(t=55, b=95, l=60, r=20),
                )
                fig_avg.update_xaxes(
                    showgrid=True,
                    gridcolor=_PLT_GRID,
                    linecolor=_PLT_LINE,
                    tickangle=-25,
                    automargin=True,
                    tickfont=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                    title_font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                    title_standoff=18,
                )
                fig_avg.update_yaxes(
                    showgrid=True,
                    gridcolor=_PLT_GRID,
                    linecolor=_PLT_LINE,
                    automargin=True,
                    tickfont=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                    title_font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                    title_standoff=10,
                )
                st.plotly_chart(fig_avg, use_container_width=True, theme=None)

            # Advanced insight charts
            st.markdown("#### Advanced Insights")

            def _depth_to_m(label):
                if label is None:
                    return np.nan
                s = str(label)
                if "5-10" in s:
                    return 0.05
                m = re.search(r"(\d+(?:[.,]\d+)?)\s*m", s)
                if m:
                    return float(m.group(1).replace(",", "."))
                return np.nan

            # 1) Depth-Time Heatmap (Hovmoller) for one point
            hov_point = st.selectbox(
                "Depth-Time point",
                options=sel_points,
                format_func=lambda x: f"Σημείο {x}",
                key="ts_hov_point",
            )
            # Use all depths for Hovmoller so the chart remains available
            # even when the main depth filter is set to a single level.
            hov_source = df[df["point"].isin(sel_points)].copy()
            hov_df = hov_source[hov_source["point"] == hov_point][["date", "depth", param]].copy()
            hov_df["depth_m"] = hov_df["depth"].apply(_depth_to_m)
            hov_df = hov_df.dropna(subset=["depth_m", param])
            if hov_df["date"].nunique() >= 2 and hov_df["depth_m"].nunique() >= 2:
                hov_pivot = (
                    hov_df.groupby(["depth_m", "date"])[param]
                    .mean()
                    .unstack("date")
                    .sort_index()
                )
                all_dates = sorted(pd.to_datetime(filt["date"]).dropna().unique())
                if all_dates:
                    hov_pivot = hov_pivot.reindex(columns=all_dates)

                fig_hov = px.imshow(
                    hov_pivot,
                    labels=dict(x="Date", y="Depth (m)", color=param),
                    color_continuous_scale="RdYlBu_r",
                    aspect="auto",
                    title=f"Depth-Time Heatmap - Point {hov_point} ({param})",
                    height=360,
                )
                fig_hov.update_layout(
                    plot_bgcolor=_PLT_BG,
                    paper_bgcolor=_PLT_PAPER,
                    font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                    margin=dict(t=55, b=25, l=45, r=20),
                )
                fig_hov.update_xaxes(
                    tickformat="%d/%m/%Y",
                    tickangle=-30,
                    showgrid=True,
                    gridcolor=_PLT_GRID,
                    linecolor=_PLT_LINE,
                    tickfont=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                    title_font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                )
                fig_hov.update_yaxes(
                    autorange="reversed",
                    showgrid=True,
                    gridcolor=_PLT_GRID,
                    linecolor=_PLT_LINE,
                    tickfont=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                    title_font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                )
                st.plotly_chart(fig_hov, use_container_width=True, theme=None)
                if sel_depth != "Όλα":
                    st.caption("Το Hovmöller εμφανίζεται με όλα τα βάθη για να υπάρχει πλήρης κατακόρυφη πληροφορία.")
            else:
                st.info("Depth-Time Heatmap needs at least 2 dates and 2 depth levels for the selected point.")

            # 2) Stratification time series (surface - deep)
            temp_col = next((c for c in df.columns if "°C" in c or "°C" in c), None)
            do_col = next((c for c in df.columns if "DO" in c), None)
            strat_cols = [c for c in [temp_col, do_col] if c]
            if strat_cols:
                strat_df = filt[["date", "point", "depth"] + strat_cols].copy()
                strat_df["depth_m"] = strat_df["depth"].apply(_depth_to_m)
                strat_df = strat_df.dropna(subset=["depth_m"])

                if not strat_df.empty:
                    surf = (
                        strat_df[strat_df["depth_m"] <= 1.0]
                        .groupby(["date", "point"])[strat_cols]
                        .mean()
                        .reset_index()
                    )
                    deep_idx = strat_df.groupby(["date", "point"])["depth_m"].idxmax()
                    deep = strat_df.loc[deep_idx, ["date", "point"] + strat_cols].copy()
                    deep = deep.rename(columns={c: f"{c}_deep" for c in strat_cols})
                    strat_merged = surf.merge(deep, on=["date", "point"], how="inner")

                    metric_cols = []
                    if temp_col and f"{temp_col}_deep" in strat_merged.columns:
                        strat_merged["delta_temp"] = strat_merged[temp_col] - strat_merged[f"{temp_col}_deep"]
                        metric_cols.append("delta_temp")
                    if do_col and f"{do_col}_deep" in strat_merged.columns:
                        strat_merged["delta_do"] = strat_merged[do_col] - strat_merged[f"{do_col}_deep"]
                        metric_cols.append("delta_do")

                    if metric_cols:
                        strat_ts = strat_merged.groupby("date")[metric_cols].mean().reset_index().sort_values("date")
                        if not strat_ts.empty:
                            fig_strat = make_subplots(specs=[[{"secondary_y": True}]])
                            if "delta_temp" in strat_ts.columns:
                                fig_strat.add_trace(
                                    go.Scatter(
                                        x=strat_ts["date"],
                                        y=strat_ts["delta_temp"],
                                        mode="lines+markers",
                                        name="ΔTemp (Surface-Deep)",
                                        line=dict(color="#f97316", width=2.5),
                                    ),
                                    secondary_y=False,
                                )
                            if "delta_do" in strat_ts.columns:
                                fig_strat.add_trace(
                                    go.Scatter(
                                        x=strat_ts["date"],
                                        y=strat_ts["delta_do"],
                                        mode="lines+markers",
                                        name="ΔDO (Surface-Deep)",
                                        line=dict(color=_PLT_SERIES[1], width=2.5),
                                    ),
                                    secondary_y=True,
                                )
                            fig_strat.update_layout(
                                title="Stratification Metrics Over Time",
                                height=360,
                                plot_bgcolor=_PLT_BG,
                                paper_bgcolor=_PLT_PAPER,
                                font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                                legend=dict(
                                    orientation="h",
                                    yanchor="bottom",
                                    y=1.02,
                                    xanchor="left",
                                    x=0,
                                    bgcolor=_PLT_LEGEND_BG,
                                    bordercolor=_PLT_LEGEND_BORDER,
                                    borderwidth=1,
                                ),
                                margin=dict(t=55, b=25, l=45, r=45),
                            )
                            fig_strat.update_xaxes(title="Date", tickformat="%d/%m/%Y", tickangle=-30, showgrid=True, gridcolor=_PLT_GRID)
                            fig_strat.update_yaxes(title_text="ΔTemp (°C)", showgrid=True, gridcolor=_PLT_GRID, secondary_y=False)
                            fig_strat.update_yaxes(title_text="ΔDO (mg/L)", showgrid=False, secondary_y=True)
                            fig_strat.update_xaxes(
                                showgrid=True,
                                gridcolor=_PLT_GRID,
                                linecolor=_PLT_LINE,
                                tickfont=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                                title_font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                            )
                            fig_strat.update_yaxes(
                                showgrid=True,
                                gridcolor=_PLT_GRID,
                                linecolor=_PLT_LINE,
                                tickfont=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                                title_font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                                secondary_y=False,
                            )
                            fig_strat.update_yaxes(
                                linecolor=_PLT_LINE,
                                tickfont=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                                title_font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                                secondary_y=True,
                            )
                            st.plotly_chart(fig_strat, use_container_width=True, theme=None)

            # 3) Anomaly chart (z-score vs global baseline)
            global_vals = pd.to_numeric(df[param], errors="coerce").dropna()
            if len(global_vals) > 1 and float(global_vals.std(ddof=0)) > 0:
                anom_ts = grp_nonnull.groupby("date")[param].mean().reset_index().sort_values("date")
                mu = float(global_vals.mean())
                sigma = float(global_vals.std(ddof=0))
                anom_ts["z"] = (anom_ts[param] - mu) / sigma
                bar_colors = ["#ef4444" if abs(v) >= 2 else _PLT_SERIES[1] for v in anom_ts["z"]]

                fig_anom = go.Figure()
                fig_anom.add_trace(go.Bar(
                    x=anom_ts["date"],
                    y=anom_ts["z"],
                    marker_color=bar_colors,
                    name="z-score",
                    hovertemplate="Date: %{x|%d/%m/%Y}<br>z-score: %{y:.2f}<extra></extra>",
                ))
                fig_anom.add_hline(y=2, line_dash="dash", line_color="#ef4444")
                fig_anom.add_hline(y=-2, line_dash="dash", line_color="#ef4444")
                fig_anom.add_hline(y=0, line_dash="dot", line_color="#64748b")
                fig_anom.update_layout(
                    title=f"Anomaly (z-score) for {param} - Global Baseline",
                    xaxis=dict(title="Date", tickformat="%d/%m/%Y", tickangle=-30),
                    yaxis=dict(title="z-score"),
                    height=320,
                    plot_bgcolor=_PLT_BG,
                    paper_bgcolor=_PLT_PAPER,
                    font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                    margin=dict(t=55, b=25, l=45, r=20),
                )
                fig_anom.update_xaxes(
                    showgrid=True,
                    gridcolor=_PLT_GRID,
                    linecolor=_PLT_LINE,
                    tickfont=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                    title_font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                )
                fig_anom.update_yaxes(
                    showgrid=True,
                    gridcolor=_PLT_GRID,
                    linecolor=_PLT_LINE,
                    tickfont=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                    title_font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                )
                st.plotly_chart(fig_anom, use_container_width=True, theme=None)

# ══════════════════════════════════════════════════════════════
# TAB 4: 3D MAPS
# ══════════════════════════════════════════════════════════════
with tab_3d:
    st.subheader("Unified 3D Georeferenced Map")
    st.caption(f"Data source path: `{MEASUREMENTS_SOURCE_PATH}`")

    c3d_a, c3d_b, c3d_c = st.columns(1) if _is_mobile else st.columns([2, 2, 2])
    with c3d_a:
        param_3d = st.selectbox(
            "Παράμετρος 3D",
            options=[p for p in COL_MAP if df[p].notna().any()],
            key="map3d_param",
        )
    with c3d_b:
        all_points_3d = sorted(df["point"].unique())
        sel_points_3d = st.multiselect(
            "Σημεία Δειγματοληψίας",
            options=all_points_3d,
            default=all_points_3d,
            format_func=lambda x: f"Σημείο {x}",
            key="map3d_points",
        )
    with c3d_c:
        all_depths_3d = [d for d in DEPTH_ORDER if d in df["depth"].unique()]
        sel_depth_3d = st.selectbox(
            "Βάθος 3D",
            options=["Όλα"] + all_depths_3d,
            key="map3d_depth",
        )

    if not sel_points_3d:
        st.warning("Επιλέξτε τουλάχιστον ένα σημείο για τον 3D χάρτη.")
    else:
        filt_3d = df[df["point"].isin(sel_points_3d)].copy()
        if sel_depth_3d != "Όλα":
            filt_3d = filt_3d[filt_3d["depth"] == sel_depth_3d]

        valid_3d = filt_3d[["date", "point", "depth", param_3d]].copy()
        valid_3d[param_3d] = pd.to_numeric(valid_3d[param_3d], errors="coerce")
        valid_3d["depth_m"] = valid_3d["depth"].apply(depth_to_m)
        valid_3d = valid_3d.dropna(subset=["date", "point", "depth_m", param_3d]).sort_values("date")

        if valid_3d.empty:
            st.info("No 3D georeferenced data is available for the current filters.")
        else:
            all_dates_valid = sorted(pd.to_datetime(valid_3d["date"]).dropna().unique())
            mode_3d = st.radio(
                "3D Χρονικό mode",
                options=["Single date", "All dates"],
                horizontal=True,
                key="map3d_mode",
            )
            ui_c1, ui_c2, ui_c3 = st.columns([1.2, 1, 1])
            with ui_c1:
                map_theme_3d = st.selectbox(
                    "Map theme",
                    options=["Voyager", "Positron", "Dark Matter"],
                    key="map3d_theme",
                )
            with ui_c2:
                depth_scale_3d = st.slider(
                    "Depth scale",
                    min_value=0.6,
                    max_value=2.4,
                    value=1.2,
                    step=0.1,
                    key="map3d_depth_scale",
                )
            with ui_c3:
                show_labels_3d = st.checkbox(
                    "Show station labels",
                    value=True,
                    key="map3d_labels",
                )

            if mode_3d == "Single date":
                sel_map_date_3d = st.select_slider(
                    "Χρονική στιγμή 3D χάρτη",
                    options=all_dates_valid,
                    value=all_dates_valid[-1],
                    format_func=lambda d: pd.to_datetime(d).strftime("%d/%m/%Y"),
                    key="map3d_lake_date",
                )
                sel_date_norm = pd.to_datetime(sel_map_date_3d).normalize()
                date_series_norm = pd.to_datetime(valid_3d["date"], errors="coerce").dt.normalize()
                map_3d_frame = valid_3d[date_series_norm == sel_date_norm].copy()
            else:
                map_3d_frame = valid_3d.copy()

            deck_obj, pmin, pmax = build_3d_lake_deck(
                map_3d_frame,
                param_3d,
                pitch=58,
                bearing=28,
                zoom_override=None,
                depth_exaggeration=depth_scale_3d,
                map_theme=map_theme_3d,
                show_labels=show_labels_3d,
                drag_mode="rotate",
            )
            if deck_obj is None:
                st.info("3D map has no points for the selected mode/date.")
            else:
                st.pydeck_chart(deck_obj, use_container_width=True)
                rendered_n = int(len(map_3d_frame.dropna(subset=[param_3d])))
                st.caption(
                    f"{param_3d}: global min {pmin:.3f} έως global max {pmax:.3f}. "
                    f"Αποδόθηκαν {rendered_n} μετρήσεις. "
                    "Mouse drag για περιστροφή, scroll για zoom, διπλό-click για focus. "
                    "Το ίδιο georeferenced 3D map δείχνει θέση (lat/lon), βάθος (z) και χρόνο (slider ή all dates)."
                )

# ══════════════════════════════════════════════════════════════
# TAB 5: DEPTH PROFILES
# ══════════════════════════════════════════════════════════════
with tab_depth:
    st.subheader("Κατακόρυφα Προφίλ")
    
    col_x, col_y, col_z = st.columns(3)
    with col_x:
        default_dp_dates = [dates[-1]] if len(dates) > 0 else []
        sel_dates_dp = st.multiselect(
            "Ημερομηνίες",
            options=dates,
            default=default_dp_dates,
            format_func=lambda d: d.strftime("%d/%m/%Y"),
            key="dp_dates"
        )
    with col_y:
        dp_point_options = ["All points"] + sorted(df["point"].unique())
        sel_point_dp = st.selectbox(
            "Σημείο",
            options=dp_point_options,
            format_func=lambda x: x if isinstance(x, str) else f"Σημείο {x}",
            key="dp_point_v2"
        )
    with col_z:
        profile_params = st.multiselect(
            "Παράμετροι",
            options=[p for p in COL_MAP if df[p].notna().any()],
            default=["pH", "Θερμοκρασία (°C)", "Διαλυμένο Οξυγόνο DO (mg/L)"],
            key="dp_params"
        )
    
    if profile_params:
        if not sel_dates_dp:
            st.info("Επίλεξε τουλάχιστον μία ημερομηνία για να εμφανιστούν τα κατακόρυφα προφίλ.")
        else:
            # Map depth to numeric value
            depth_numeric = {
                "Επιφάνεια (5-10 cm)": 0.05, "Βάθος 1m": 1, "Βάθος 5m": 5,
                "Βάθος 10m": 10, "Βάθος 12m": 12, "Βάθος 15m": 15,
                "Βάθος 20m": 20, "Βάθος 23m": 23, "Βάθος 25m": 25
            }
            dp_colors = px.colors.qualitative.Plotly
            dates_sorted = sorted(sel_dates_dp)
            range_source = df[df["date"].isin(dates_sorted)].copy()
            if sel_point_dp != "All points":
                range_source = range_source[range_source["point"] == sel_point_dp]
            range_source["depth_m"] = range_source["depth"].map(depth_numeric)
            range_source = range_source.dropna(subset=["depth_m"])
            param_axis_ranges = {}
            depth_axis_range = None
            for param in profile_params:
                vals = pd.to_numeric(range_source[param], errors="coerce").dropna()
                if vals.empty:
                    continue
                pmin = float(vals.min())
                pmax = float(vals.max())
                span = pmax - pmin
                pad = span * 0.06 if span > 0 else max(abs(pmin) * 0.05, 0.1)
                param_axis_ranges[param] = [pmin - pad, pmax + pad]
            depth_vals = pd.to_numeric(range_source["depth_m"], errors="coerce").dropna()
            if not depth_vals.empty:
                dmin = float(depth_vals.min())
                dmax = float(depth_vals.max())
                dspan = dmax - dmin
                dpad = dspan * 0.04 if dspan > 0 else max(abs(dmax) * 0.05, 0.5)
                depth_axis_range = [dmax + dpad, max(0.0, dmin - dpad)]
            
            for row_idx, sel_date_dp in enumerate(dates_sorted):
                sub = df[df["date"] == sel_date_dp].copy()
                if sel_point_dp != "All points":
                    sub = sub[sub["point"] == sel_point_dp]
                sub["depth_m"] = sub["depth"].map(depth_numeric)
                sub = sub.dropna(subset=["depth_m"]).sort_values("depth_m")
                
                st.markdown(f"#### Ημερομηνία: {sel_date_dp.strftime('%d/%m/%Y')}")
                if sub.empty:
                    st.info("Δεν υπάρχουν δεδομένα για αυτή τη συνδυαστική επιλογή.")
                    continue
                
                n_cols = max(1, len(profile_params))
                cols = st.columns(n_cols)
                
                for i, param in enumerate(profile_params):
                    fig_p = go.Figure()
                    if sel_point_dp == "All points":
                        point_list = sorted(sub["point"].dropna().unique())
                        for idx_pt, pt in enumerate(point_list):
                            psub = sub[sub["point"] == pt][["depth_m", param]].dropna().sort_values("depth_m")
                            if psub.empty:
                                continue
                            clr = dp_colors[idx_pt % len(dp_colors)]
                            fig_p.add_trace(go.Scatter(
                                x=psub[param], y=psub["depth_m"],
                                mode="lines+markers",
                                name=f"Σημείο {int(pt)}",
                                line=dict(color=clr, width=2.0),
                                marker=dict(size=6, color=clr, line=dict(width=1.0, color="#ffffff")),
                            ))
                    else:
                        psub = sub[["depth_m", param]].dropna().sort_values("depth_m")
                        if psub.empty:
                            continue
                        fig_p.add_trace(go.Scatter(
                            x=psub[param], y=psub["depth_m"],
                            mode="lines+markers",
                            name=f"Σημείο {sel_point_dp}",
                            line=dict(color="#2e86c1", width=2.1),
                            marker=dict(size=7, color="#1a5276", line=dict(width=1.2, color="#ffffff"))
                        ))
                    
                    if len(fig_p.data) == 0:
                        continue
                    fig_p.update_layout(
                        xaxis=dict(side="bottom"),
                        yaxis=dict(title="Βάθος (m)", autorange="reversed"),
                        height=380,
                        plot_bgcolor="#ffffff",
                        paper_bgcolor="#ffffff",
                        font=dict(color="#334155", family="Plus Jakarta Sans, sans-serif"),
                        legend=dict(
                            orientation="v",
                            yanchor="bottom",
                            y=0.03,
                            xanchor="right",
                            x=0.98,
                            font=dict(size=10, color="#64748b"),
                            bgcolor="rgba(255,255,255,0.82)",
                            bordercolor="#e2e8f0",
                            borderwidth=1
                        ),
                        hoverlabel=dict(
                            bgcolor="#ffffff",
                            font_color="#0f172a",
                            bordercolor="#cbd5e1"
                        ),
                        margin=dict(t=10, b=56, l=52, r=12)
                    )
                    fig_p.update_xaxes(
                        title_text=param,
                        side="bottom",
                        range=param_axis_ranges.get(param),
                        showgrid=True,
                        gridcolor="rgba(148,163,184,0.22)",
                        linecolor="rgba(148,163,184,0.5)",
                        tickfont=dict(color="#334155", family="Plus Jakarta Sans, sans-serif"),
                        title_font=dict(color="#0f172a"),
                        title_standoff=12,
                        automargin=True
                    )
                    fig_p.update_yaxes(
                        range=depth_axis_range,
                        showgrid=True,
                        gridcolor="rgba(148,163,184,0.22)",
                        linecolor="rgba(148,163,184,0.5)",
                        tickfont=dict(color="#334155", family="Plus Jakarta Sans, sans-serif"),
                        title_font=dict(color="#0f172a"),
                        automargin=True
                    )
                    with cols[i % n_cols]:
                        st.plotly_chart(
                            fig_p,
                            use_container_width=True,
                            theme=None,
                            config={"displayModeBar": False},
                        )
                
                if row_idx < len(dates_sorted) - 1:
                    st.markdown("---")

# ══════════════════════════════════════════════════════════════
# TAB 4: COMPARE POINTS
# ══════════════════════════════════════════════════════════════
with tab_compare:
    st.subheader("Σύγκριση Σημείων – Συγκεκριμένη Ημερομηνία")
    
    col1c, col2c = st.columns(2)
    with col1c:
        cmp_date = st.selectbox(
            "Ημερομηνία",
            options=dates,
            format_func=lambda d: d.strftime("%d/%m/%Y"),
            key="cmp_date"
        )
    with col2c:
        cmp_depth = st.selectbox(
            "Βάθος",
            options=["Μέσος Όρος"] + [d for d in DEPTH_ORDER if d in df["depth"].unique()],
            key="cmp_depth"
        )
    
    sub_cmp = df[df["date"] == cmp_date].copy()
    if cmp_depth != "Μέσος Όρος":
        sub_cmp = sub_cmp[sub_cmp["depth"] == cmp_depth]
    
    grp_cmp = sub_cmp.groupby("point")[[p for p in COL_MAP if df[p].notna().any()]].mean().reset_index()
    
    avail_params_cmp = [p for p in COL_MAP if p in grp_cmp.columns and grp_cmp[p].notna().any()]
    
    if not avail_params_cmp:
        st.info("Δεν υπάρχουν δεδομένα.")
    else:
        param_cmp = st.selectbox("Παράμετρος για σύγκριση", avail_params_cmp, key="cmp_param")
        
        plot_cmp = grp_cmp[["point", param_cmp]].dropna()
        
        if not plot_cmp.empty:
            fig_cmp = px.bar(
                plot_cmp,
                x="point", y=param_cmp,
                labels={"point": "Σημείο Δειγματοληψίας", param_cmp: param_cmp},
                color=param_cmp,
                color_continuous_scale="Blues",
                text_auto=".2f",
                title=f"{param_cmp} ανά Σημείο — {cmp_date.strftime('%d/%m/%Y')}",
                height=420
            )
            fig_cmp.update_layout(
                xaxis=dict(tickmode="array", tickvals=plot_cmp["point"], ticktext=[f"Σ{p}" for p in plot_cmp["point"]]),
                plot_bgcolor=_PLT_BG,
                paper_bgcolor=_PLT_PAPER,
                font=dict(color=_PLT_TICK)
            )
            fig_cmp.update_xaxes(
                showgrid=True,
                gridcolor=_PLT_GRID,
                linecolor=_PLT_LINE,
                tickfont=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                title_font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
            )
            fig_cmp.update_yaxes(
                showgrid=True,
                gridcolor=_PLT_GRID,
                linecolor=_PLT_LINE,
                tickfont=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
                title_font=dict(color=_PLT_TICK, family="Plus Jakarta Sans, sans-serif"),
            )
            st.plotly_chart(fig_cmp, use_container_width=True, theme=None)
        
        # Radar chart for all params at this date
        st.markdown("#### 🕸️ Ραδιογράφημα – Όλα τα Σημεία")
        radar_params = [p for p in ["pH", "Θερμοκρασία (°C)", "Διαλυμένο Οξυγόνο DO (mg/L)",
                                     "Αγωγιμότητα (μS/cm)", "Θολότητα-Εργαστήριο (NTU)", "Χλωροφύλλη-α (μg/L)"]
                        if p in grp_cmp.columns and grp_cmp[p].notna().any()]
        
        if radar_params:
            fig_r = go.Figure()
            colors_r = _PLT_SERIES
            for idx, row_r in grp_cmp.iterrows():
                vals = [row_r[p] for p in radar_params]
                if all(np.isnan(v) for v in vals):
                    continue
                # Normalize each param
                normed = []
                for p_idx, p in enumerate(radar_params):
                    col_vals = grp_cmp[p].dropna()
                    mn, mx = col_vals.min(), col_vals.max()
                    v = row_r[p]
                    normed.append((v - mn) / (mx - mn) if mx != mn and not np.isnan(v) else 0)
                
                pt_idx = int(row_r['point']) % len(colors_r)
                clr = colors_r[pt_idx]
                m_rgb = re.match(r"rgba?\(\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)", str(clr))
                if m_rgb:
                    r_c, g_c, b_c = map(int, m_rgb.groups())
                else:
                    clr_hex = str(clr).lstrip("#")
                    if len(clr_hex) == 3:
                        clr_hex = "".join(ch * 2 for ch in clr_hex)
                    try:
                        r_c = int(clr_hex[0:2], 16)
                        g_c = int(clr_hex[2:4], 16)
                        b_c = int(clr_hex[4:6], 16)
                    except Exception:
                        r_c, g_c, b_c = 46, 134, 193
                        clr = f"rgb({r_c},{g_c},{b_c})"
                fig_r.add_trace(go.Scatterpolar(
                    r=normed + [normed[0]],
                    theta=radar_params + [radar_params[0]],
                    fill="toself",
                    name=f"Σημείο {int(row_r['point'])}",
                    line=dict(color=clr),
                    fillcolor=f"rgba({r_c},{g_c},{b_c},0.12)"
                ))
            fig_r.update_layout(
                polar=dict(radialaxis=dict(visible=True, range=[0, 1], tickformat=".0%")),
                height=420,
                title=f"Κανονικοποιημένο Ραδιογράφημα — {cmp_date.strftime('%d/%m/%Y')}"
            )
            st.plotly_chart(fig_r, use_container_width=True, theme=None)

# ══════════════════════════════════════════════════════════════
# TAB 5: RAW DATA
# ══════════════════════════════════════════════════════════════
with tab_raw:
    st.subheader("Ακατέργαστα Δεδομένα")
    
    c1, c2 = st.columns(2)
    with c1:
        filter_points = st.multiselect(
            "Φίλτρο Σημείων",
            options=sorted(df["point"].unique()),
            default=list(sorted(df["point"].unique())),
            format_func=lambda x: f"Σημείο {x}",
            key="raw_pts"
        )
    with c2:
        filter_dates = st.multiselect(
            "Φίλτρο Ημερομηνιών",
            options=dates,
            format_func=lambda d: d.strftime("%d/%m/%Y"),
            default=dates[-5:],
            key="raw_dates"
        )
    
    filt_raw = df[df["point"].isin(filter_points) & df["date"].isin(filter_dates)].copy()
    filt_raw["date"] = filt_raw["date"].dt.strftime("%d/%m/%Y")
    filt_raw = filt_raw.rename(columns={"date": "Ημερομηνία", "point": "Σημείο", "depth": "Βάθος"})
    
    st.dataframe(filt_raw.reset_index(drop=True), use_container_width=True, height=400)
    
    csv = filt_raw.to_csv(index=False).encode("utf-8-sig")
    st.download_button("⬇️ Λήψη CSV", data=csv, file_name="monitoring_data.csv", mime="text/csv")

st.markdown("---")
st.caption("ΕΥΑΘ · Δορυφορική Παρακολούθηση Φράγματος Γαδουρά · 2025-2026")
