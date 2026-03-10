from PIL import Image
import os
import re
import pytesseract
from openpyxl import Workbook, load_workbook
from datetime import datetime
import cv2
import numpy as np
from pathlib import Path

# Set the Tesseract executable path
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# Function to preprocess the image
def preprocess_image(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    # Apply adaptive thresholding for better OCR performance
    thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                                   cv2.THRESH_BINARY_INV, 11, 2)
    return thresh

# Function to extract date from image
def extract_date(image_path):
    img = cv2.imdecode(np.fromfile(str(image_path), dtype=np.uint8), cv2.IMREAD_COLOR)
    if img is None:
        print(f"Error: Unable to load image at {image_path}")
        return None
    try:
        # Get image dimensions
        height, width, _ = img.shape

        # Define the region of interest (top-right corner) dynamically
        roi_top = 0
        roi_bottom = int(0.10 * height)  # Top 10% of the image
        roi_left = int(0.80 * width)     # Rightmost 20% of the image
        roi_right = width

        # First attempt: Dynamic cropping
        date_area = img[roi_top:roi_bottom, roi_left:roi_right]

        # Preprocess and perform OCR
        preprocessed = preprocess_image(date_area)
        date_text = pytesseract.image_to_string(preprocessed, config='--psm 7')

        # Debugging: Print the raw OCR text
        print(f"Dynamic OCR text from {image_path}: '{date_text.strip()}'")

        # Try parsing the extracted date
        formatted_date = re.findall(r'\d{4}-\d{2}-\d{2}', date_text)
        if formatted_date:
            return datetime.strptime(formatted_date[0], '%Y-%m-%d')
        
        # Fallback: Fixed cropping if dynamic fails
        print(f"Dynamic OCR failed, attempting fixed cropping for {image_path}")
        date_area_fixed = img[0:50, 362:512]
        preprocessed_fixed = preprocess_image(date_area_fixed)
        date_text_fixed = pytesseract.image_to_string(preprocessed_fixed, config='--psm 7')

        # Debugging: Print the fixed OCR text
        print(f"Fixed OCR text from {image_path}: '{date_text_fixed.strip()}'")

        # Try parsing the date from the fixed cropping
        formatted_date_fixed = re.findall(r'\d{4}-\d{2}-\d{2}', date_text_fixed)
        if formatted_date_fixed:
            return datetime.strptime(formatted_date_fixed[0], '%Y-%m-%d')
        else:
            print(f"Date not found in both dynamic and fixed attempts for {image_path}")
            return None
    except Exception as e:
        print(f"Error processing image {image_path}: {e}")
        return None

# Function to extract number from filename
def extract_number(filename):
    number = re.findall(r'\d+$', filename.stem)
    return int(number[0]) if number else 0

# Use current directory for processing
folder_path = Path('.')

# Check if the folder exists
if not folder_path.exists():
    print("Error: The target folder does not exist.")
    exit(1)

# Read images and extract data
data = []

for filename in sorted(folder_path.iterdir()):
    if filename.suffix.lower() in ('.png', '.jpg', '.jpeg'):
        print(f"Processing file: {filename}")
        date_obj = extract_date(filename)
        file_number = extract_number(filename)
        if date_obj:
            data.append((file_number, date_obj.year, date_obj.month, date_obj.day))
        else:
            data.append((file_number, None, None, None))

# Function to handle sorting logic for None values
def sort_key(row):
    return tuple(x if x is not None else float('inf') for x in row)

# Sort data before writing to Excel, taking care of None values
sorted_data = sorted(data, key=sort_key)

# Save data to Excel file
output_path = folder_path / 'ocr.xlsx'

wb = Workbook()
ws = wb.active

# Write the numbering, year, month, and day to the output sheet
ws.cell(row=1, column=1, value='File Number')
ws.cell(row=1, column=2, value='Year')
ws.cell(row=1, column=3, value='Month')
ws.cell(row=1, column=4, value='Day')

for row_idx, row_data in enumerate(sorted_data, start=2):
    for col_idx, cell_data in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=cell_data)

wb.save(str(output_path))

# Read data from Excel file for further processing if needed
wb = load_workbook(str(output_path))
ws = wb.active

data = []
for row in ws.iter_rows(min_row=1, values_only=True):
    data.append(row)

# Overwrite the existing Excel file with the sorted data if needed